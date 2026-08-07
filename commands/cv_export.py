import asyncio
import sqlite3
import io
import datetime
from hasad_bot.config import config
from hasad_bot.utils import now_hijri
from telegram.ext import Application
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
async def export_cv():
    app = Application.builder().token(config.bot_token).build()
    channel_id = config.backup_channel_id
    conn = sqlite3.connect(config.harvest_db)
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM student_cvs_v2')
    rows = cursor.fetchall()
    if not rows:
        print('⚠️ لا توجد بيانات')
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'سجلات الطلاب'
    ws.sheet_view.rightToLeft = True
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    bold_font = Font(bold=True, color='FFFFFF')
    headers = ['Telegram ID', 'Platform User', 'الاسم بالعربية', 'الاسم باللاتينية', 'رقم الهوية', 'الجوال', 'الجنسية', 'المرحلة', 'الصف', 'الفصل', 'تاريخ السحب']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        scrape_date = ''
        if len(row) > 11 and row[11]:
            scrape_date = datetime.datetime.fromtimestamp(row[11]).strftime('%Y-%m-%d %H:%M')
        ws.append([row[1], row[0], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], scrape_date])
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    now = datetime.datetime.now()
    filename = f'CV_Export_{now.strftime("%Y%m%d_%H%M")}.xlsx'
    excel_stream.name = filename
    await app.bot.send_document(chat_id=channel_id, document=excel_stream, caption=f'🚨 **تم استخراج سجلات الطلاب** 🕵️\n📅 {now_hijri()}', parse_mode='HTML')
    print(f'✅ تم إرسال {len(rows)} سجل للقناة')
asyncio.run(export_cv())
