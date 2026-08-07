import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import asyncio
import io
import datetime
import os
import pyzipper
from hasad_bot.database import db_all_users
from hasad_bot.utils import decrypt_password, now_hijri
from hasad_bot.config import config
from telegram.ext import Application
from telegram import InputFile
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

async def export_creds():
    app = Application.builder().token(config.bot_token).build()
    channel_id = config.backup_channel_id
    
    # ✅ كلمة المرور من ملف .env
    zip_password = os.environ.get("EXCEL_PASSWORD", os.environ.get("ZIP_PASSWORD", "Hasad_Default_2024")).encode()
    
    users = await db_all_users()
    
    filtered = []
    for u in users:
        if u.get('dars360_user') and u.get('dars360_pass'):
            pass_plain = decrypt_password(u['dars360_pass'])
            filtered.append({
                'telegram_id': u['telegram_id'], 
                'name': u.get('name', ''), 
                'platform_user': u['dars360_user'], 
                'password': pass_plain, 
                'expiry': u.get('expiry_hijri', ''), 
                'free_attempts': u.get('free_attempts', 0)
            })
    
    if not filtered:
        print('⚠️ لا توجد بيانات')
        return
    
    # ✅ إنشاء ملف Excel مع تنسيقات محسنة
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'بيانات المنصة'
    ws.sheet_view.rightToLeft = True
    
    # تنسيقات
    header_fill = PatternFill(start_color='8B0000', end_color='8B0000', fill_type='solid')
    bold_font = Font(bold=True, color='FFFFFF', size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ✅ إضافة الهيدر
    headers = ['Telegram ID', 'الاسم', 'يوزر المنصة', 'كلمة المرور', 'الصلاحية', 'محاولات مجانية', 'VIP']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # ✅ إضافة البيانات
    vip_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
    alt_row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    for idx, u in enumerate(filtered):
        is_vip = u['expiry'] and u['expiry'] not in ['', 'تم الإلغاء ❌']
        
        row_num = idx + 2
        ws.append([u['telegram_id'], u['name'], u['platform_user'], u['password'], u['expiry'], u['free_attempts'], '✅' if is_vip else '❌'])
        
        # تنسيق الصف
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.alignment = center_align
            cell.border = thin_border
            
            if idx % 2 == 0:
                cell.fill = alt_row_fill
            
            if is_vip:
                cell.fill = vip_fill
            
            if col == 4:
                cell.font = Font(name='Courier New', size=10, bold=True)
    
    # ✅ ضبط عرض الأعمدة
    column_widths = [15, 25, 20, 30, 20, 15, 8]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[chr(65 + i)].width = width
    
    # ✅ حفظ Excel في الذاكرة
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    excel_filename = f"Platform_Credentials_{timestamp}.xlsx"
    zip_filename = f"Platform_Credentials_{timestamp}.zip"
    
    # ✅ إنشاء ZIP مشفر يحتوي على ملف Excel
    with pyzipper.AESZipFile(
        zip_filename, 
        'w',
        compression=pyzipper.ZIP_DEFLATED,
        encryption=pyzipper.WZ_AES
    ) as zipf:
        zipf.setpassword(zip_password)
        zipf.writestr(excel_filename, excel_stream.getvalue())
    
    # ✅ إرسال الملف المشفر
    with open(zip_filename, 'rb') as doc:
        await app.bot.send_document(
            chat_id=channel_id, 
            document=InputFile(doc, filename=zip_filename),
            caption=(
                f'<b>🔑 استخراج بيانات المنصة - مشفر</b> 🔑\n\n'
                f'<b>📊 عدد الحسابات:</b> {len(filtered)}\n'
                f'<b>✅ VIP:</b> {sum(1 for u in filtered if u["expiry"] and u["expiry"] not in ["", "تم الإلغاء ❌"])}\n'
                f'<b>📅 التاريخ:</b> {now_hijri()}\n'
                f'<b>🔒 نوع التشفير:</b> AES-256\n'
                f'<b>📁 حجم الملف:</b> {os.path.getsize(zip_filename) / 1024:.1f} KB'
            ),
            parse_mode='HTML'
        )
    
    # ✅ إرسال كلمة المرور في رسالة منفصلة
    await app.bot.send_message(
        chat_id=channel_id,
        text=(
            f'<b>🔑 كلمة مرور فك الضغط</b> 🔑\n\n'
            f'⚠️ <b>هذا الملف يحتوي على معلومات حساسة</b>\n'
            f'📌 احتفظ بكلمة المرور في مكان آمن'
        ),
        parse_mode='HTML'
    )
    
    # ✅ حذف الملف المحلي
    os.remove(zip_filename)
    
    vip_count = sum(1 for u in filtered if u["expiry"] and u["expiry"] not in ["", "تم الإلغاء ❌"])
    print(f'✅ تم إرسال {len(filtered)} حساب (مشفر) للقناة | VIP: {vip_count} | كلمة المرور: {zip_password.decode()}')

asyncio.run(export_creds())