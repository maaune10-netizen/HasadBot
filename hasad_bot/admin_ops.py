# -*- coding: utf-8 -*-
"""
عمليات الإدارة المشتركة - تُستخدم من Telegram (main.py) ومن لوحة التحكم (Web Dashboard)
إرسال الملفات المشفرة، النسخ الاحتياطي لقاعدة البيانات، وتصدير بيانات المنصة
"""

import asyncio
import io
import math
import os
import re
import tempfile
import time
import zipfile
import datetime as dt
from typing import Tuple, Dict, List, Optional
from uuid import uuid4
import aiosqlite
import msoffcrypto
import pyzipper
from pathlib import Path

from loguru import logger

from hasad_bot.config import config
from hasad_bot.utils import now_hijri, admin_trace, decrypt_password, gregorian_to_hijri
from hasad_bot.datetime_utils import datetime, now_timestamp
from hasad_bot.database import (
    _db_pool,
    db_all_users,
    db_get_user,
    db_set_user,
    db_delete_user,
    db_log,
    create_user_subscription,
    update_user_free_attempts,
    archive_user_credentials,
    log_admin_action,
    get_users_by_target,
    get_users_count_by_target,
    is_admin,
    is_reseller,
    get_all_full_admins,
    get_admin_sub_resellers,
    promote_to_admin,
    demote_from_admin,
    promote_to_reseller,
    demote_from_reseller,
    add_reseller_credit,
    get_reseller_credit,
    get_reseller_stats,
    set_reseller_credit_price,
    get_all_reseller_credit_prices,
    is_bot_frozen,
    set_bot_frozen,
    is_public_mode,
    set_public_mode,
)


# ==============================================================================
# حارس الأثر الخارجي (Side-effect guard)
# ==============================================================================

class OperationBlocked(Exception):
    """رفض تنفيذ عملية ذات أثر خارجي (بيئة اختبار غير معزولة)"""
    pass


def assert_side_effect_safe() -> None:
    # بيئة اختبار بدون data dir معزول = ممنوع أي عملية لها أثر خارجي
    if config.app_env == "test" and not config.allow_live_tests:
        if not os.environ.get("HASAD_DATA_DIR"):
            raise OperationBlocked("APP_ENV=test بدون HASAD_DATA_DIR معزول — ممنوع تنفيذ عمليات ذات أثر خارجي")


async def send_encrypted_excel_file(bot, chat_id, workbook, filename: str, caption: str):
    """تشفير Excel وإرساله (يفتح على الجوال)"""
    assert_side_effect_safe()
    password = config.backup_password
    
    temp_input = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_input.close()
    workbook.save(temp_input.name)
    
    temp_output = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_output.close()
    
    try:
        with open(temp_input.name, 'rb') as f:
            office_file = msoffcrypto.OfficeFile(f)
            with open(temp_output.name, 'wb') as f_out:
                office_file.encrypt(password, f_out)
        
        with open(temp_output.name, 'rb') as f:
            await bot.send_document(
                chat_id=chat_id,
                document=f,
                filename=filename,
                caption=f"{caption}\n🔐 **كلمة المرور:** `{password}`\n📅 {now_hijri()}\n\n✅ يفتح على Excel مباشرة (جوال + كمبيوتر)",
                parse_mode="Markdown"
            )
    finally:
        if os.path.exists(temp_input.name):
            os.unlink(temp_input.name)
        if os.path.exists(temp_output.name):
            os.unlink(temp_output.name)


async def send_encrypted_zip_file(bot, chat_id, file_path, caption: str):
    """إرسال ZIP مشفر (لقاعدة البيانات واللوجات)"""
    assert_side_effect_safe()
    from pathlib import Path
    
    password = config.backup_password
    original_name = file_path.name
    zip_filename = f"{original_name}.zip"
    zip_path = Path(tempfile.gettempdir()) / zip_filename
    
    try:
        with pyzipper.AESZipFile(zip_path, 'w', compression=pyzipper.ZIP_DEFLATED, encryption=pyzipper.WZ_AES) as zipf:
            zipf.setpassword(password.encode())
            zipf.write(file_path, original_name)
        
        with open(zip_path, 'rb') as f:
            await bot.send_document(
                chat_id=chat_id,
                document=f,
                filename=zip_filename,
                caption=f"{caption}\n🔐 **كلمة المرور:** `{password}`\n📅 {now_hijri()}\n\n⚠️ ملف ZIP محمي (يفتح على الكمبيوتر)",
                parse_mode="Markdown"
            )
    finally:
        if zip_path.exists():
            zip_path.unlink()


async def send_db_backup(bot, chat_id=None):
    """Send database backup to channel"""
    assert_side_effect_safe()
    db_path = config.knowledge_db
    channel_id = chat_id if chat_id is not None else config.backup_channel_id
    
    if not os.path.exists(db_path):
        logger.warning("⚠️ No database to backup")
        print("⚠️ لا توجد قاعدة بيانات للنسخ الاحتياطي")
        return
    
    if not channel_id:
        logger.warning("⚠️ BACKUP_CHANNEL_ID not set")
        print("⚠️ BACKUP_CHANNEL_ID غير معرف في ملف .env")
        return
    
    zip_filename = f"Hasad_DB_Backup_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    ok = False
    
    try:
        with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(db_path, os.path.basename(db_path))
        
        with open(zip_filename, 'rb') as doc:
            await bot.send_document(
                chat_id=channel_id,
                document=doc,
                caption=f"📦 **قاعدة المعرفة**\n📅 {now_hijri()}",
                parse_mode="Markdown"
            )
        ok = True
        
        admin_trace("BACKUP_SUCCESS", "Database backup sent")
        logger.success("✅ Backup sent")
        print(f"✅ تم إرسال النسخة الاحتياطية للقناة: {zip_filename}")
        
    except Exception as e:
        logger.error(f"Backup error: {e}")
        print(f"❌ خطأ في النسخ الاحتياطي: {e}")
    finally:
        if os.path.exists(zip_filename):
            os.remove(zip_filename)
    return ok


async def send_cv_export(bot, chat_id=None):
    """Export CV data to Excel - مع تنسيقات كيو كيو"""
    assert_side_effect_safe()
    print("✅ بدء تصدير بيانات CV...")
    
    channel_id = chat_id if chat_id is not None else config.backup_channel_id
    
    if not channel_id:
        print("❌ BACKUP_CHANNEL_ID غير معرف في ملف .env")
        return
    
    if not config.harvest_db.exists():
        print(f"❌ ملف قاعدة البيانات غير موجود: {config.harvest_db}")
        return
    ok = False
    
    try:
        async with aiosqlite.connect(config.harvest_db) as db:
            # إنشاء الجدول إذا لم يكن موجوداً
            await db.execute('''
                CREATE TABLE IF NOT EXISTS student_cvs_v2 (
                    platform_user TEXT PRIMARY KEY,
                    telegram_id INTEGER,
                    local_name TEXT,
                    latin_name TEXT,
                    identity_no TEXT,
                    phone TEXT,
                    nationality TEXT,
                    stage TEXT,
                    grade TEXT,
                    student_class TEXT,
                    profile_pic TEXT,
                    scraped_at REAL
                )
            ''')
            
            # عد السجلات
            async with db.execute("SELECT COUNT(*) FROM student_cvs_v2") as cursor:
                count = await cursor.fetchone()
                total_records = count[0] if count else 0
                print(f"📊 عدد السجلات: {total_records}")
            
            if total_records == 0:
                print("❌ لا توجد بيانات في الجدول")
                return
            
            # إنشاء ملف Excel بالتنسيقات
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "سجلات الطلاب (حصاد)"
            ws.sheet_view.rightToLeft = True
            
            # تنسيقات كيو كيو
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            green_fill  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            bold_white_font = Font(name='Tahoma', size=11, bold=True, color="FFFFFF")
            normal_font = Font(name='Tahoma', size=10)
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            headers = [
                'Telegram ID', 'Platform User', 'الاسم بالعربية', 'الاسم باللاتينية',
                'رقم الهوية', 'الجوال', 'الجنسية', 'المرحلة', 'الصف', 'الفصل', 'تاريخ السحب'
            ]
            ws.append(headers)
            
            # تنسيق الهيدر
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = bold_white_font
                cell.alignment = center_align
                cell.border = thin_border
            
            rows_added = 0
            async with db.execute("SELECT * FROM student_cvs_v2") as cursor:
                async for row in cursor:
                    scrape_date = dt.datetime.fromtimestamp(row[11]).strftime('%Y-%m-%d %H:%M') if len(row) > 11 else ""
                    row_data = [
                        row[1], row[0], row[2], row[3], row[4],
                        row[5], row[6], row[7], row[8], row[9], scrape_date
                    ]
                    ws.append(row_data)
                    rows_added += 1
                    
                    # تنسيق الصفوف مثل كيو كيو
                    current_row = ws[ws.max_row]
                    for col_idx, cell in enumerate(current_row):
                        cell.font = normal_font
                        cell.alignment = center_align
                        cell.border = thin_border
                        
                        # تلوين الصفوف الفردية والزوجية
                        if rows_added % 2 == 0:
                            cell.fill = alt_row_fill
                        
                        # تلوين خاص لبعض الأعمدة
                        if col_idx == 1:  # Platform User
                            cell.fill = yellow_fill
                            cell.font = Font(name='Tahoma', size=10, bold=True, color="C00000")
                        elif col_idx == 10:  # تاريخ السحب
                            cell.fill = green_fill
            
            # ضبط عرض الأعمدة
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min((max_length + 4), 40)
            
            # حفظ الملف
            excel_stream = io.BytesIO()
            wb.save(excel_stream)
            excel_stream.seek(0)
            excel_stream.name = f"CV_Export_{dt.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            
            print(f"📁 تم إنشاء الملف: {excel_stream.name}")
            print(f"📤 جاري الإرسال للقناة...")
            
            # إرسال للقناة
            caption = f"🚨 <b>تم استخراج سجلات الطلاب</b> 🕵️\nعدد السجلات: <b>{rows_added}</b>\n📅 {now_hijri()}"
            
            await bot.send_document(
                chat_id=channel_id,
                document=excel_stream,
                caption=caption,
                parse_mode="HTML"
            )
            ok = True
            
            print(f"✅ تم إرسال ملف الإكسل الملون للقناة! ({rows_added} سجل)")
            
    except Exception as e:
        logger.error(f"CV export error: {e}")
        print(f"❌ خطأ في تصدير الملف: {e}")
        import traceback
        traceback.print_exc()
    return ok


async def extract_credentials(bot, chat_id=None):
    """استخراج بيانات المنصة - ملف Excel ملون ومدلع"""
    assert_side_effect_safe()
    print("🔑 بدء استخراج بيانات المنصة...")
    
    channel_id = chat_id if chat_id is not None else config.backup_channel_id
    
    if not channel_id:
        print("❌ BACKUP_CHANNEL_ID غير معرف في ملف .env")
        return
    ok = False
    
    try:
        users = await db_all_users()
        
        # فلترة المستخدمين اللي عندهم حسابات منصة
        filtered_users = []
        for u in users:
            if u.get('dars360_user') and u.get('dars360_pass'):
                pass_plain = decrypt_password(u['dars360_pass'])
                filtered_users.append({
                    'telegram_id': u['telegram_id'],
                    'name': u.get('name', ''),
                    'platform_user': u['dars360_user'],
                    'password': pass_plain,
                    'expiry': u.get('expiry_hijri', ''),
                    'free_attempts': u.get('free_attempts', 0)
                })
        
        if not filtered_users:
            print("📭 لا توجد بيانات منصة مخزنة.")
            return
        
        print(f"📊 عدد الحسابات: {len(filtered_users)}")
        
        # إنشاء ملف Excel بالتنسيقات
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        import io
        import datetime
        
        wb = Workbook()
        ws = wb.active
        ws.title = "بيانات المنصة"
        ws.sheet_view.rightToLeft = True
        
        # تنسيقات كيو كيو
        header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # أحمر غامق
        vip_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # أخضر فاتح للـ VIP
        normal_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        bold_white_font = Font(name='Tahoma', size=11, bold=True, color="FFFFFF")
        normal_font = Font(name='Tahoma', size=10)
        bold_red_font = Font(name='Tahoma', size=10, bold=True, color="8B0000")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        headers = [
            'Telegram ID', 'الاسم', 'يوزر المنصة', 'كلمة المرور (مفكوكة)',
            'الصلاحية', 'واجبات مجانية', 'VIP'
        ]
        ws.append(headers)
        
        # تنسيق الهيدر
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = bold_white_font
            cell.alignment = center_align
            cell.border = thin_border
        
        # إضافة البيانات
        for idx, u in enumerate(filtered_users):
            # التحقق إذا كان المستخدم VIP (مشترك)
            is_vip = u['expiry'] and u['expiry'] not in ['', 'تم الإلغاء ❌']
            
            row_data = [
                u['telegram_id'],
                u['name'],
                u['platform_user'],
                u['password'],
                u['expiry'],
                u['free_attempts'],
                '✅' if is_vip else '❌'
            ]
            ws.append(row_data)
            
            # تنسيق الصف
            current_row = ws[ws.max_row]
            for col_idx, cell in enumerate(current_row):
                cell.font = normal_font
                cell.alignment = center_align
                cell.border = thin_border
                
                # تلوين الصفوف الفردية والزوجية
                if idx % 2 == 0:
                    cell.fill = alt_row_fill
                
                # تلوين خاص للأعمدة
                if col_idx == 2:  # يوزر المنصة
                    cell.font = bold_red_font
                elif col_idx == 3:  # كلمة المرور
                    cell.font = Font(name='Courier New', size=10, bold=True)
                
                # تلوين صفوف VIP بالكامل
                if is_vip:
                    cell.fill = vip_fill
        
        # ضبط عرض الأعمدة
        column_widths = [15, 25, 20, 25, 20, 15, 8]
        for i, width in enumerate(column_widths):
            ws.column_dimensions[chr(65 + i)].width = width
        
        # حفظ الملف
        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        excel_stream.name = f"Platform_Credentials_{dt.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        print(f"📁 تم إنشاء الملف: {excel_stream.name}")
        print(f"📤 جاري الإرسال للقناة...")
        
        # إرسال للقناة
        caption = f"🔑 <b>استخراج بيانات المنصة</b>\n"
        caption += f"عدد الحسابات: <b>{len(filtered_users)}</b>\n"
        caption += f"✅ VIP: <b>{sum(1 for u in filtered_users if u['expiry'] and u['expiry'] not in ['', 'تم الإلغاء ❌'])}</b>\n"
        caption += f"📅 {now_hijri()}"
        
        await bot.send_document(
            chat_id=channel_id,
            document=excel_stream,
            caption=caption,
            parse_mode="HTML"
        )
        ok = True
        
        print(f"✅ تم إرسال ملف البيانات للقناة! ({len(filtered_users)} حساب)")
        
        # كمان نعرض في التيرمينال ملخص
        print(f"\n🔑 ملخص الحسابات المستخرجة:")
        print(f"   إجمالي: {len(filtered_users)}")
        print(f"   VIP: {sum(1 for u in filtered_users if u['expiry'] and u['expiry'] not in ['', 'تم الإلغاء ❌'])}")
        print(f"   عادي: {sum(1 for u in filtered_users if not u['expiry'] or u['expiry'] in ['', 'تم الإلغاء ❌'])}")
        
    except Exception as e:
        logger.error(f"Extract error: {e}")
        print(f"❌ خطأ في الاستخراج: {e}")
        import traceback
        traceback.print_exc()
    return ok


# اسم قديم محتفظ به للتوافق مع الاستدعاءات القائمة (terminal/CLI)
extract_credentials_terminal = extract_credentials


async def send_encrypted_file(bot, chat_id: int, file_path: Path, caption: str, password: str = None, custom_name: str = None):
    """
    إرسال ملف مشفر ومحمي بكلمة مرور
    """
    assert_side_effect_safe()
    import tempfile
    import os
    import pyzipper
    
    if password is None:
        password = config.backup_password
    
    # ✅ تحديد اسم الملف النهائي
    if custom_name:
        zip_filename = custom_name
    else:
        # استخراج اسم الملف الأصلي بدون مسار
        original_name = file_path.stem
        zip_filename = f"{original_name}.zip"
    
    # إنشاء ملف ZIP مؤقت باسم منظم
    temp_dir = tempfile.gettempdir()
    zip_path = Path(temp_dir) / zip_filename
    
    try:
        # إنشاء ZIP مشفر بكلمة مرور
        with pyzipper.AESZipFile(zip_path, 'w', compression=pyzipper.ZIP_DEFLATED, encryption=pyzipper.WZ_AES) as zipf:
            zipf.setpassword(password.encode())
            zipf.write(file_path, os.path.basename(file_path))
        
        # إرسال الملف
        with open(zip_path, 'rb') as f:
            await bot.send_document(
                chat_id=chat_id,
                document=f,
                filename=zip_filename,  # ✅ اسم واضح ومنظم
                caption=f"🔒 **{caption}**\n🔐 **محمي بكلمة مرور**\n📅 {now_hijri()}\n\n📌 **كلمة المرور:** `{password}`",
                parse_mode="Markdown"
            )
            
    finally:
        # حذف الملف المؤقت
        if zip_path.exists():
            zip_path.unlink()


# ==============================================================================
# عمليات الإدارة المشتركة - تُستخدم من Telegram ومن لوحة التحكم (Web Dashboard)
# كل دالة: تنفيذ التعديلات على قاعدة البيانات + إشعار المستخدم + تسجيل العملية
# ==============================================================================

async def _get_payment_request_by_id(request_id: int):
    """قراءة سجل طلب دفع بالمعرّف (None إذا لم يوجد)"""
    try:
        conn = await _db_pool.get_connection()
        async with conn.execute(
            "SELECT * FROM payment_requests WHERE id = ?",
            (request_id,)
        ) as cursor:
            row = await cursor.fetchone()
        return dict(row) if row else None
    except Exception as e:
        logger.error(f"Error reading payment request {request_id}: {e}")
        return None


async def renew_subscription(bot, uid: int, days: int, actor: str = "dashboard") -> Tuple[bool, str]:
    """تجديد اشتراك مستخدم (نفس منطق admin_renew_got_days)"""
    assert_side_effect_safe()
    try:
        u = await db_get_user(uid)
        if not u:
            return (False, "❌ المستخدم غير موجود.")

        cur_exp = u.get("expiry_ts", 0) or 0
        if cur_exp < now_timestamp():
            cur_exp = now_timestamp()

        new_exp = cur_exp + days * 86400
        exp_h = gregorian_to_hijri(datetime.fromtimestamp(new_exp))
        # تحديث جدول users
        await db_set_user(uid, expiry_ts=new_exp, expiry_hijri=exp_h)

        # تحديد الخطة حسب عدد الأيام
        if days <= 7:
            plan_id = "weekly"
        elif days <= 30:
            plan_id = "monthly"
        else:
            plan_id = "semester"

        await create_user_subscription(uid, plan_id, cur_exp, new_exp)

        # إشعار للمستخدم
        try:
            await bot.send_message(
                uid,
                f"🎉 <b>تم تحديث اشتراكك!</b> +{days} يوم.\nالانتهاء: {exp_h}",
                parse_mode="HTML"
            )
        except Exception:
            pass

        await log_admin_action(0, actor, "RENEW_SUBSCRIPTION",
                               target_user_id=uid, target_user_name=u.get("name", ""),
                               old_value=str(u.get("expiry_ts", 0) or 0), new_value=str(new_exp),
                               details=f"+{days} days, plan={plan_id}, end={exp_h}")
        admin_trace("RENEW_SUBSCRIPTION", f"User {uid} +{days}d (plan={plan_id}) by {actor}", uid=str(uid))

        return (True, f"✅ تم تجديد <code>{uid}</code> +{days} يوم | الانتهاء: {exp_h}")

    except Exception as e:
        logger.error(f"renew_subscription error: {e}")
        return (False, f"❌ خطأ: {e}")


async def revoke_subscription(bot, uid: int, actor: str = "dashboard") -> Tuple[bool, str]:
    """إلغاء اشتراك مستخدم (نفس منطق admin_revoke_done)"""
    assert_side_effect_safe()
    try:
        u = await db_get_user(uid)
        if not u:
            return (False, "❌ المستخدم غير موجود.")

        await db_set_user(uid, expiry_ts=0, expiry_hijri="تم الإلغاء ❌")

        # إشعار للمستخدم
        try:
            await bot.send_message(
                uid,
                "🚫 <b>تم إلغاء اشتراكك من قبل الإدارة.</b>",
                parse_mode="HTML"
            )
        except Exception:
            pass

        await log_admin_action(0, actor, "REVOKE_SUBSCRIPTION",
                               target_user_id=uid, target_user_name=u.get("name", ""),
                               old_value=str(u.get("expiry_ts", 0) or 0), new_value="0",
                               details="Revoked by " + actor)
        admin_trace("REVOKE_SUBSCRIPTION", f"User {uid} revoked by {actor}", uid=str(uid))

        return (True, f"✅ تم إلغاء اشتراك <code>{uid}</code>.")

    except Exception as e:
        logger.error(f"revoke_subscription error: {e}")
        return (False, f"❌ خطأ: {e}")


async def add_homework_credit(bot, uid: int, count: int, kind: str) -> Tuple[bool, str]:
    """إضافة واجبات لمستخدم (kind: free للرصيد المجاني، sub لحد الاشتراك)"""
    assert_side_effect_safe()
    try:
        user = await db_get_user(uid)
        if not user:
            return (False, "❌ المستخدم غير موجود.")
        name = user.get("name", uid)

        if kind == "free":
            # إضافة إلى free_attempts
            current = user.get("free_attempts", 0)
            new_value = current + count
            await update_user_free_attempts(uid, new_value)

            await db_log(0, "ADD_HOMEWORKS",
                         detail=f"User {uid} +{count} free (was {current}, now {new_value})")

            # إشعار للمستخدم
            try:
                await bot.send_message(
                    uid,
                    f"🎉 <b>تم إضافة {count} واجبات مجانية إلى رصيدك!</b>\n"
                    f"🎟️ رصيدك الحالي: {new_value} واجب",
                    parse_mode="HTML"
                )
            except Exception:
                pass

            await log_admin_action(0, "admin", "ADD_HOMEWORKS_FREE",
                                   target_user_id=uid, target_user_name=name,
                                   old_value=str(current), new_value=str(new_value),
                                   details=f"+{count} free homeworks")
            admin_trace("ADD_HOMEWORKS", f"User {uid} +{count} free (was {current}, now {new_value})", uid=str(uid))

            return (True, f"✅ <b>تمت الإضافة بنجاح!</b>\n\n"
                          f"👤 المستخدم: {name}\n"
                          f"➕ تم إضافة: {count} واجب (رصيد مجاني)\n"
                          f"🎟️ الرصيد الجديد: {new_value}")

        elif kind == "sub":
            # إضافة إلى max_homeworks في الاشتراك النشط
            conn = await _db_pool.get_connection()
            async with conn.execute("""
                SELECT id, max_homeworks FROM user_subscriptions
                WHERE user_id = ? AND is_active = 1 AND end_date > ?
                ORDER BY end_date DESC LIMIT 1
            """, (uid, time.time())) as cursor:
                sub = await cursor.fetchone()
            if not sub:
                return (False, "❌ لا يوجد اشتراك نشط لهذا المستخدم.")
            sub_id, current_max = sub
            new_max = current_max + count
            await conn.execute(
                "UPDATE user_subscriptions SET max_homeworks = ? WHERE id = ?",
                (new_max, sub_id)
            )
            await conn.commit()

            await db_log(0, "ADD_HOMEWORKS_SUB",
                         detail=f"User {uid} +{count} to subscription (was {current_max}, now {new_max})")

            # إشعار للمستخدم
            try:
                await bot.send_message(
                    uid,
                    f"🎉 <b>تم زيادة حد اشتراكك بمقدار {count} واجب!</b>\n"
                    f"📦 الحد الجديد: {new_max} واجب",
                    parse_mode="HTML"
                )
            except Exception:
                pass

            await log_admin_action(0, "admin", "ADD_HOMEWORKS_SUB",
                                   target_user_id=uid, target_user_name=name,
                                   old_value=str(current_max), new_value=str(new_max),
                                   details=f"+{count} to subscription limit")
            admin_trace("ADD_HOMEWORKS_SUB", f"User {uid} +{count} to subscription (was {current_max}, now {new_max})", uid=str(uid))

            return (True, f"✅ تم إضافة {count} واجبات إلى حد اشتراك المستخدم {name}.\n"
                          f"📦 الحد الجديد: {new_max} واجب")

        else:
            return (False, f"❌ نوع غير معروف: {kind}")

    except Exception as e:
        logger.error(f"add_homework_credit error: {e}")
        return (False, f"❌ خطأ: {e}")


async def approve_payment_request(bot, request_id: int, days: int, actor: str = "dashboard") -> Tuple[bool, str]:
    """تفعيل اشتراك من طلب دفع (نفس منطق set_days_callback / handle_custom_days_input)"""
    assert_side_effect_safe()
    try:
        req = await _get_payment_request_by_id(request_id)
        if not req:
            return (False, "❌ طلب الدفع غير موجود.")
        if req.get("status") != "pending":
            return (False, f"❌ الطلب رقم {request_id} تمت معالجته مسبقاً (الحالة: {req.get('status')}).")

        uid = req["user_id"]
        u = await db_get_user(uid)
        if not u:
            return (False, f"❌ المستخدم {uid} غير موجود")

        # حساب تواريخ الاشتراك
        cur_exp = u.get("expiry_ts", 0) or 0
        if cur_exp < now_timestamp():
            cur_exp = now_timestamp()

        new_exp = cur_exp + days * 86400
        exp_h = gregorian_to_hijri(datetime.fromtimestamp(new_exp))
        # تحديث جدول users
        await db_set_user(uid, expiry_ts=new_exp, expiry_hijri=exp_h)

        # تحديد الخطة حسب عدد الأيام
        if days <= 7:
            plan_id = "weekly"
            max_homeworks = 25
        elif days <= 30:
            plan_id = "monthly"
            max_homeworks = 100
        else:
            plan_id = "semester"
            max_homeworks = 200

        await create_user_subscription(uid, plan_id, cur_exp, new_exp)

        # تحديث طلب الدفع
        try:
            conn = await _db_pool.get_connection()
            await conn.execute("""
                UPDATE payment_requests
                SET status = 'approved', processed_at = ?, processed_by = ?
                WHERE id = ? AND status = 'pending'
            """, (time.time(), actor, request_id))
            await conn.commit()
        except Exception:
            pass

        # إشعار للمستخدم
        try:
            await bot.send_message(
                uid,
                f"🎉 <b>تم تفعيل اشتراكك!</b> 🎉\n\n"
                f"📦 <b>الخطة:</b> {plan_id}\n"
                f"📚 <b>الواجبات:</b> {max_homeworks} واجب\n"
                f"📅 <b>المدة:</b> +{days} يوم\n"
                f"📆 <b>الانتهاء:</b> {exp_h}\n\n"
                f"🚀 استمتع بحل الواجبات!",
                parse_mode="HTML"
            )
        except Exception:
            pass

        await log_admin_action(0, actor, "APPROVE_PAYMENT_REQUEST",
                               target_user_id=uid, target_user_name=req.get("user_name", ""),
                               old_value="pending", new_value="approved",
                               details=f"Request #{request_id}, +{days} days, plan={plan_id}")
        admin_trace("APPROVE_PAYMENT_REQUEST", f"Request #{request_id} approved for user {uid} (+{days}d, plan={plan_id}) by {actor}", uid=str(uid))

        return (True, f"✅ <b>تم التفعيل بنجاح!</b>\n\n"
                      f"👤 المستخدم: <code>{uid}</code>\n"
                      f"📅 المدة: {days} يوم\n"
                      f"📦 الخطة: {plan_id} ({max_homeworks} واجب)\n"
                      f"📆 الانتهاء: {exp_h}")

    except Exception as e:
        logger.error(f"approve_payment_request error: {e}")
        return (False, f"❌ خطأ: {e}")


async def reject_payment_request(bot, request_id: int, reason: str, actor: str = "dashboard") -> Tuple[bool, str]:
    """رفض طلب دفع (نفس منطق reject_reason_callback / handle_custom_reject)"""
    assert_side_effect_safe()
    try:
        req = await _get_payment_request_by_id(request_id)
        if not req:
            return (False, "❌ طلب الدفع غير موجود.")
        if req.get("status") != "pending":
            return (False, f"❌ الطلب رقم {request_id} تمت معالجته مسبقاً (الحالة: {req.get('status')}).")

        uid = req["user_id"]

        # تحديث طلب الدفع
        try:
            conn = await _db_pool.get_connection()
            await conn.execute("""
                UPDATE payment_requests
                SET status = 'rejected', processed_at = ?, processed_by = ?
                WHERE id = ? AND status = 'pending'
            """, (time.time(), actor, request_id))
            await conn.commit()
        except Exception:
            pass

        # إشعار للمستخدم
        try:
            await bot.send_message(
                uid,
                f"❌ <b>تم رفض طلب الاشتراك</b>\n\n"
                f"📌 <b>السبب:</b> {reason}\n\n"
                f"💡 للاستفسار، تواصل مع الدعم الفني.",
                parse_mode="HTML"
            )
        except Exception:
            pass

        await log_admin_action(0, actor, "REJECT_PAYMENT_REQUEST",
                               target_user_id=uid, target_user_name=req.get("user_name", ""),
                               old_value="pending", new_value="rejected",
                               details=f"Request #{request_id}, reason: {reason}")
        admin_trace("REJECT_PAYMENT_REQUEST", f"Request #{request_id} rejected for user {uid} by {actor}", uid=str(uid))

        return (True, f"✅ <b>تم رفض الطلب!</b>\n\n"
                      f"👤 المستخدم: <code>{uid}</code>\n"
                      f"📌 السبب: {reason}")

    except Exception as e:
        logger.error(f"reject_payment_request error: {e}")
        return (False, f"❌ خطأ: {e}")


async def unlock_user(bot, uid: int, actor: str = "dashboard") -> Tuple[bool, str]:
    """فك قفل مستخدم مع أرشفة بيانات المنصة إن وجدت (نفس منطق cb_unlock)"""
    assert_side_effect_safe()
    try:
        u = await db_get_user(uid)
        if not u:
            return (False, "❌ المستخدم غير موجود.")

        # أرشفة بيانات المنصة قبل فك القفل (إن وجدت)
        try:
            await archive_user_credentials(uid, 0, actor, "فك القفل بواسطة الإدارة")
        except Exception:
            pass

        await db_set_user(uid, locked_to=None, lock_request=0, dars360_user=None, dars360_pass=None)

        # إشعار للمستخدم
        try:
            await bot.send_message(
                uid,
                "🔓 <b>تم فك قفل حسابك.</b> يمكنك إضافة حساب جديد الآن.",
                parse_mode="HTML"
            )
        except Exception:
            pass

        await log_admin_action(0, actor, "UNLOCK_USER",
                               target_user_id=uid, target_user_name=u.get("name", ""),
                               details="Unlocked by " + actor)
        admin_trace("UNLOCK_USER", f"User {uid} unlocked by {actor}", uid=str(uid))

        return (True, f"✅ تم فك قفل <code>{uid}</code>.")

    except Exception as e:
        logger.error(f"unlock_user error: {e}")
        return (False, f"❌ خطأ: {e}")


async def delete_user(bot, uid: int, actor: str = "dashboard") -> Tuple[bool, str]:
    """حذف مستخدم مع جميع سجلاته (نفس منطق cb_delete)"""
    assert_side_effect_safe()
    try:
        u = await db_get_user(uid)
        if not u:
            return (False, "❌ المستخدم غير موجود.")

        await db_delete_user(uid)

        await log_admin_action(0, actor, "DELETE_USER",
                               target_user_id=uid, target_user_name=u.get("name", ""),
                               details="Deleted by " + actor)
        admin_trace("DELETE_USER", f"User {uid} deleted by {actor}", uid=str(uid))

        return (True, f"🗑️ تم حذف <code>{uid}</code>.")

    except Exception as e:
        logger.error(f"delete_user error: {e}")
        return (False, f"❌ خطأ: {e}")


# ==============================================================================
# الإرسال الجماعي (Broadcast) والإعلانات — منطق مشترك مع لوحة التحكم
# ==============================================================================

# مخزن وظائف الإرسال الحية (broadcast + announcements)
SEND_JOBS: Dict[str, dict] = {}


def get_send_job(job_id: str) -> Optional[dict]:
    """قراءة نسخة آمنة من بيانات وظيفة إرسال (نسخة وليس مرجع)."""
    now = time.time()
    # تنظيف الوظائف المنتهية منذ أكثر من ساعة لمنع نمو SEND_JOBS بلا حدود
    for jid in [jid for jid, j in SEND_JOBS.items()
                if j.get("status") == "done" and (j.get("finished_at") or 0) < now - 3600]:
        SEND_JOBS.pop(jid, None)
    # حد أقصى: إذا تجاوز العدد 200 وظيفة، احذف أقدم الوظائف المنتهية
    if len(SEND_JOBS) > 200:
        done_jobs = sorted(
            (jid for jid, j in SEND_JOBS.items() if j.get("status") == "done"),
            key=lambda jid: SEND_JOBS[jid].get("finished_at") or 0,
        )
        for jid in done_jobs[: len(SEND_JOBS) - 200]:
            SEND_JOBS.pop(jid, None)

    job = SEND_JOBS.get(job_id)
    if job is None:
        return None
    return {k: (list(v) if isinstance(v, list) else v) for k, v in job.items()}


def _new_job_id() -> str:
    """معرّف فريد لوظيفة إرسال."""
    return uuid4().hex


async def send_broadcast(bot, target: str, text: str, actor: str = "dashboard", progress_cb=None,
                         admin_id: int = 0, admin_name: str = None) -> dict:
    """إرسال رسالة جماعية نصية لفئة مستخدمين (نفس منطق admin_broadcast_send).

    target: all | subscribed | not_subscribed | linked | not_linked
    admin_id/admin_name: هوية الأدمن لسجل التدقيق (admin_name يقع افتراضياً على actor).
    Returns: {"sent": n, "failed": n, "skipped": n, "errors": [...]}
    """
    assert_side_effect_safe()
    valid_targets = ("all", "subscribed", "not_subscribed", "linked", "not_linked")
    if target not in valid_targets:
        return {"sent": 0, "failed": 0, "skipped": 0, "errors": [f"فئة غير معروفة: {target}"]}

    users = await get_users_by_target(target)
    total = len(users)
    sent = 0
    failed = 0
    skipped = 0
    errors: List[str] = []
    message = f"📢 <b>رسالة من الإدارة</b>\n\n{text}"

    admin_trace("BROADCAST_SEND", f"target={target} total={total} by {actor}", actor)

    for user_id in users:
        try:
            await bot.send_message(chat_id=user_id, text=message, parse_mode="HTML")
            sent += 1
        except Exception as e:
            err_msg = str(e)
            low = err_msg.lower()
            if any(k in low for k in ("blocked", "deactivated", "chat not found", "bot was kicked")):
                skipped += 1
            else:
                failed += 1
                errors.append(f"UID {user_id}: {err_msg[:80]}")
                logger.error(f"Failed to send to {user_id}: {e}")

        await asyncio.sleep(0.05)  # تجنب الـ Flood wait

        if progress_cb:
            progress_cb({
                "done": sent + failed + skipped,
                "total": total,
                "sent": sent,
                "failed": failed,
                "skipped": skipped,
            })

    await log_admin_action(admin_id, admin_name or actor, "BROADCAST_SENT",
                           details=f"target={target} total={total} sent={sent} failed={failed} skipped={skipped}")
    admin_trace("BROADCAST_DONE", f"target={target} total={total} sent={sent} failed={failed} skipped={skipped}", actor)

    return {"sent": sent, "failed": failed, "skipped": skipped, "errors": errors}


def _update_job_progress(job_id: str, p: dict):
    """تحديث عدادات الوظيفة أثناء الإرسال."""
    job = SEND_JOBS.get(job_id)
    if not job:
        return
    job["sent"] = p.get("sent", 0)
    job["failed"] = p.get("failed", 0)
    job["skipped"] = p.get("skipped", 0)
    job["total"] = p.get("total", job["total"])


async def start_broadcast(bot, target: str, text: str, actor: str = "dashboard",
                          admin_id: int = 0, admin_name: str = None) -> str:
    """إنشاء وظيفة إرسال جماعي وتشغيلها في الخلفية. يرجع job_id."""
    assert_side_effect_safe()
    job_id = _new_job_id()
    total = await get_users_count_by_target(target)
    SEND_JOBS[job_id] = {
        "job_id": job_id,
        "kind": "broadcast",
        "label": target,
        "status": "queued",
        "total": total,
        "sent": 0,
        "failed": 0,
        "skipped": 0,
        "errors": [],
        "started_at": time.time(),
        "finished_at": None,
    }
    asyncio.create_task(_run_broadcast_job(job_id, bot, target, text, actor, admin_id, admin_name))
    return job_id


async def _run_broadcast_job(job_id: str, bot, target: str, text: str, actor: str,
                             admin_id: int = 0, admin_name: str = None):
    """تنفيذ وظيفة الإرسال الجماعي وتحديث الحالة."""
    job = SEND_JOBS[job_id]
    job["status"] = "running"
    try:
        result = await send_broadcast(
            bot, target, text, actor=actor,
            progress_cb=lambda p: _update_job_progress(job_id, p),
            admin_id=admin_id, admin_name=admin_name,
        )
        job["sent"] = result["sent"]
        job["failed"] = result["failed"]
        job["skipped"] = result["skipped"]
        job["errors"] = result["errors"]
    except Exception as e:
        job["errors"].append(str(e)[:200])
        logger.error(f"broadcast job {job_id} error: {e}")
    finally:
        job["status"] = "done"
        job["finished_at"] = time.time()


async def start_announcement_send(bot, atype: str, actor: str = "dashboard") -> str:
    """إنشاء وظيفة إرسال إعلان وتشغيلها في الخلفية. يرجع job_id."""
    assert_side_effect_safe()
    job_id = _new_job_id()
    SEND_JOBS[job_id] = {
        "job_id": job_id,
        "kind": "announcement",
        "label": atype,
        "status": "queued",
        "total": 0,
        "sent": 0,
        "failed": 0,
        "skipped": 0,
        "errors": [],
        "started_at": time.time(),
        "finished_at": None,
    }
    asyncio.create_task(_run_announcement_job(job_id, bot, atype, actor))
    return job_id


async def _run_announcement_job(job_id: str, bot, atype: str, actor: str):
    """تنفيذ وظيفة إرسال الإعلان وتحديث الحالة + تسجيل العملية بعد الاكتمال."""
    from hasad_bot.handlers.announcements import send_announcement

    job = SEND_JOBS[job_id]
    job["status"] = "running"
    try:

        def _cb(p: dict):
            job["sent"] = p.get("sent", 0)
            job["skipped"] = p.get("skipped", 0)
            job["total"] = p.get("total", job["total"])
            job["progress_errors"] = p.get("errors", 0)  # عدد مؤقت أثناء التقدم

        sent, skipped, errors = await send_announcement(bot, atype, manual=True, progress_cb=_cb)
        job["sent"] = sent
        job["skipped"] = skipped
        job["failed"] = len(errors)
        job["errors"] = errors
        await log_admin_action(0, actor, "ANNOUNCEMENT_SENT",
                               details=f"type={atype} sent={sent} skipped={skipped} errors={len(errors)}")
        admin_trace("ANNOUNCEMENT_SENT", f"type={atype} sent={sent} skipped={skipped} errors={len(errors)} by {actor}", actor)
    except Exception as e:
        job.setdefault("errors", [])
        if not isinstance(job["errors"], list):
            job["errors"] = []
        job["errors"].append(str(e)[:200])
        logger.error(f"announcement job {job_id} error: {e}")
    finally:
        job["status"] = "done"
        job["finished_at"] = time.time()


# ==============================================================================
# الدعم الفني (Support) — مصدر البيانات: جدول logs (نفس تدفق تيليجرام)
# ==============================================================================

async def get_support_conversations(status: str = "all", limit: int = 100, q: str = "") -> List[dict]:
    """محادثات الدعم: المستخدمون الذين لديهم سجلات SUPPORT_MSG / SUPPORT_REPLY.

    لكل مستخدم: user_id, name, last_activity_ts, last_direction (user|admin),
    msg_count, reply_count؛ الحالة المستخلصة: open إذا كان آخر اتجاه user وإلا closed.
    الترتيب: last_activity_ts تنازلياً، ثم limit. q يفلتر بالاسم أو المعرّف قبل اقتطاع limit.
    """
    conn = await _db_pool.get_connection()
    convos: Dict[int, dict] = {}
    async with conn.execute(
        "SELECT telegram_id, action, created_at FROM logs "
        "WHERE action IN ('SUPPORT_MSG','SUPPORT_REPLY') "
        "ORDER BY telegram_id ASC, created_at ASC"
    ) as cursor:
        async for row in cursor:
            uid = row[0]
            c = convos.get(uid)
            if c is None:
                c = {
                    "user_id": uid,
                    "name": "",
                    "last_activity_ts": 0,
                    "last_direction": "user",
                    "msg_count": 0,
                    "reply_count": 0,
                    "status": "open",
                }
                convos[uid] = c
            c["last_activity_ts"] = row[2]
            if row[1] == "SUPPORT_MSG":
                c["msg_count"] += 1
                c["last_direction"] = "user"
            else:
                c["reply_count"] += 1
                c["last_direction"] = "admin"

    query = q.strip().lower() if q else ""
    results: List[dict] = []
    for c in convos.values():
        c["status"] = "open" if c["last_direction"] == "user" else "closed"
        if status != "all" and status != c["status"]:
            continue
        user = await db_get_user(c["user_id"])
        c["name"] = (user or {}).get("name", "") if user else ""
        if query and query not in str(c["name"]).lower() and query not in str(c["user_id"]):
            continue
        results.append(c)

    results.sort(key=lambda c: c["last_activity_ts"], reverse=True)
    return results[: max(0, int(limit))]


async def get_support_history(user_id: int, limit: int = 50) -> List[dict]:
    """سجل محادثة دعم لمستخدم من جدول logs (من الأحدث للأقدم)."""
    conn = await _db_pool.get_connection()
    rows: List[dict] = []
    async with conn.execute(
        "SELECT action, detail, created_at FROM logs "
        "WHERE telegram_id=? AND action IN ('SUPPORT_MSG','SUPPORT_REPLY') "
        "ORDER BY created_at DESC LIMIT ?",
        (user_id, int(limit)),
    ) as cursor:
        async for row in cursor:
            rows.append({
                "ts": row[2],
                "direction": "user" if row[0] == "SUPPORT_MSG" else "admin",
                "detail": row[1],
            })
    return rows


async def send_support_reply(bot, user_id: int, text: str, actor: str = "dashboard",
                             admin_id: int = 0, admin_name: str = "dashboard") -> Tuple[bool, str]:
    """إرسال رد دعم لمستخدم (نفس رسالة تدفق تيليجرام) + تسجيل + تدقيق."""
    assert_side_effect_safe()
    try:
        await bot.send_message(
            user_id,
            f"🛡️ <b>رد من الدعم:</b>\n\n{text}",
            parse_mode="HTML",
        )
        await db_log(user_id, "SUPPORT_REPLY", detail=text, source="ADMIN")
        await log_admin_action(admin_id, admin_name, "SUPPORT_REPLY",
                               target_user_id=user_id, details=text[:200])
        admin_trace("SUPPORT_REPLY", f"Reply to user {user_id} by {actor}", uid=str(user_id))
        return (True, "تم إرسال الرد")
    except Exception as e:
        logger.error(f"send_support_reply error: {e}")
        return (False, f"❌ تعذر إرسال الرد: {e}")


# ==============================================================================
# قراءة اللوجات (مع إخفاء الأسرار) — للوحة التحكم
# ==============================================================================

# أسماء الملفات المسموح بقراءتها (اسم العرض → مسار ضمن log_dir فقط)
LOG_FILE_ALLOWLIST: Dict[str, Path] = {
    "hasad_main": config.log_dir / "hasad_main.log",
    "hasad_errors": config.log_dir / "hasad_errors.log",
    "hasad_events": config.log_dir / "hasad_events.log",
    "hasad_security": config.log_dir / "hasad_security.log",
    "hasad_performance": config.log_dir / "hasad_performance.log",
    "admin_trace": config.accounts_log,
}


def redact_log_text(text: str) -> str:
    """إخفاء الأسرار (توكنات، مفاتيح، كلمات مرور) من نص لوج."""
    out = text
    secrets = [config.bot_token, config.jwt_secret, config.backup_password,
               config.admin_password, config.dashboard_password_hash,
               config.admin_dars_user, config.admin_dars_pass]
    secrets += list(config.groq_keys) + list(config.gemini_keys)
    for s in secrets:
        if s:
            out = out.replace(s, "[REDACTED]")
    out = re.sub(r'gsk_[A-Za-z0-9]{10,}', "[REDACTED]", out)
    out = re.sub(r'AIza[A-Za-z0-9_\-]{20,}', "[REDACTED]", out)
    out = re.sub(r'Bearer\s+\S+', "[REDACTED]", out)
    out = re.sub(r'(?i)(password|pass|token|secret)\s*[=:]\s*\S+', "[REDACTED]", out)
    return out


async def read_log_file(name: str, offset: int = 0, limit: int = 200, tail: bool = False) -> dict:
    """قراءة ملف لوج مسموح مع إخفاء الأسرار (لا يُكشف المسار الكامل أبداً)."""
    path = LOG_FILE_ALLOWLIST.get(name)
    if path is None:
        raise OperationBlocked("ملف غير مسموح")
    limit = max(1, min(int(limit), 1000))
    if not path.exists():
        return {"name": name, "total_lines": 0, "lines": [], "truncated": False}
    from collections import deque
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            if tail:
                # قراءة حتى نهاية الملف: عدّ الأسطر + الاحتفاظ بآخر limit فقط (ذاكرة محدودة)
                tail_lines = deque(maxlen=limit)
                total_lines = 0
                for ln in f:
                    total_lines += 1
                    tail_lines.append(ln.rstrip("\r\n"))
                selected = list(tail_lines)
                truncated = total_lines > limit
            else:
                start = max(0, int(offset))
                selected = []
                total_lines = 0
                for ln in f:
                    total_lines += 1
                    if total_lines <= start or len(selected) >= limit:
                        continue
                    selected.append(ln.rstrip("\r\n"))
                truncated = start + len(selected) < total_lines
    except Exception as e:
        logger.error(f"read_log_file error: {e}")
        return {"name": name, "total_lines": 0, "lines": [], "truncated": False}

    return {
        "name": name,
        "total_lines": total_lines,
        "lines": [redact_log_text(ln) for ln in selected],
        "truncated": truncated,
    }


async def get_user_log(uid: int, limit: int = 100, step_filter: str = None) -> dict:
    """لوج مستخدم (نفس محلل handlers.user_log) مع إخفاء الأسرار."""
    from hasad_bot.handlers.user_log import get_user_logs
    entries = get_user_logs(uid, limit=limit, step_filter=step_filter)
    redacted = []
    for e in entries:
        item = dict(e)
        item["detail"] = redact_log_text(item.get("detail", ""))
        redacted.append(item)
    return {"uid": uid, "entries": redacted, "total": len(redacted)}


async def get_admin_audit(q: str = "", action: str = "", limit: int = 100,
                          after: float = None, before: float = None) -> List[dict]:
    """سجل تدقيق الأدمن (admin_actions) مع فلترة وإخفاء الأسرار."""
    conn = await _db_pool.get_connection()
    sql = "SELECT * FROM admin_actions WHERE 1=1"
    params: List = []
    if action:
        sql += " AND action_type LIKE ?"
        params.append(f"%{action}%")
    if q:
        sql += " AND admin_name LIKE ?"
        params.append(f"%{q}%")
    if after is not None:
        sql += " AND created_at >= ?"
        params.append(after)
    if before is not None:
        sql += " AND created_at <= ?"
        params.append(before)
    sql += " ORDER BY id DESC LIMIT ?"
    params.append(max(1, min(int(limit), 1000)))
    rows: List[dict] = []
    async with conn.execute(sql, tuple(params)) as cursor:
        async for row in cursor:
            d = dict(row)
            if d.get("details"):
                d["details"] = redact_log_text(str(d["details"]))
            rows.append(d)
    return rows


# ==============================================================================
# النسخ الاحتياطية (Backups) — مع قفل يمنع التصدير المتزامن
# ==============================================================================

_export_lock = asyncio.Lock()


async def send_audit_logs(bot, chat_id=None):
    """إرسال ZIP مشفر يحتوي admin_actions.csv (كل الصفوف) + admin_accounts_details.log (آخر 500 سطر)."""
    assert_side_effect_safe()
    import csv
    channel_id = chat_id if chat_id is not None else config.backup_channel_id
    if not channel_id:
        raise OperationBlocked("BACKUP_CHANNEL_ID غير معرف")

    zip_path = None
    ok = False
    try:
        zip_filename = f"Hasad_AdminLogs_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        zip_path = Path(tempfile.gettempdir()) / zip_filename
        with pyzipper.AESZipFile(zip_path, 'w', compression=pyzipper.ZIP_DEFLATED,
                                 encryption=pyzipper.WZ_AES) as zipf:
            zipf.setpassword(config.backup_password.encode())

            # admin_actions.csv — كل الصفوف
            csv_buffer = io.StringIO()
            writer = csv.writer(csv_buffer)
            conn = await _db_pool.get_connection()
            async with conn.execute("SELECT * FROM admin_actions ORDER BY id DESC") as cursor:
                writer.writerow([d[0] for d in cursor.description])
                async for row in cursor:
                    writer.writerow(tuple(row))
            zipf.writestr("admin_actions.csv", csv_buffer.getvalue())

            # admin_accounts_details.log — آخر 500 سطر (قراءة سطرية بذاكرة محدودة)
            if config.accounts_log.exists():
                from collections import deque
                tail_lines = deque(maxlen=500)
                with open(config.accounts_log, "r", encoding="utf-8", errors="ignore") as f:
                    for ln in f:
                        tail_lines.append(ln.rstrip("\r\n"))
                zipf.writestr("admin_accounts_details.log", "\n".join(tail_lines) + "\n")

        with open(zip_path, "rb") as f:
            await bot.send_document(
                chat_id=channel_id,
                document=f,
                filename=zip_filename,
                caption=f"📜 <b>سجل التدقيق (Admin Logs)</b>\n🔐 **كلمة المرور:** `{config.backup_password}`\n📅 {now_hijri()}\n\n⚠️ ملف ZIP محمي (يفتح على الكمبيوتر)",
                parse_mode="Markdown"
            )
        ok = True
    finally:
        if zip_path and zip_path.exists():
            zip_path.unlink()
    return ok


async def run_backup(bot, kind: str, actor: str = "dashboard",
                     admin_id: int = 0, admin_name: str = "dashboard") -> Tuple[bool, str]:
    """تشغيل نسخة احتياطية (db | cv | admin_logs) وإرسالها لقناة النسخ الاحتياطي."""
    assert_side_effect_safe()
    if kind not in ("db", "cv", "admin_logs"):
        return (False, f"❌ نوع نسخة احتياطية غير معروف: {kind}")
    if _export_lock.locked():
        return (False, "تصدير آخر قيد التنفيذ — انتظر اكتماله")
    async with _export_lock:
        try:
            if kind == "db":
                ok = await send_db_backup(bot)
            elif kind == "cv":
                ok = await send_cv_export(bot)
            else:
                ok = await send_audit_logs(bot)
        except Exception as e:
            logger.error(f"run_backup error: {e}")
            return (False, f"❌ خطأ في النسخة الاحتياطية: {e}")
        if not ok:
            return (False, "❌ فشل إرسال النسخة الاحتياطية إلى القناة")
        await log_admin_action(admin_id, admin_name, f"BACKUP_{kind.upper()}",
                               details="sent to backup channel")
        admin_trace(f"BACKUP_{kind.upper()}", f"Backup {kind} by {actor}")
        return (True, "تم إنشاء النسخة الاحتياطية وإرسالها")


# ==============================================================================
# تحكم البوت (Bot Control) — حالة التجميد والوضع العام — مشترك مع لوحة التحكم
# ==============================================================================

async def get_bot_status() -> dict:
    """حالة البوت: تجميد / وضع عام / آخر إحصائيات / حيوية البوت"""
    frozen = await is_bot_frozen()
    public_mode = await is_public_mode()
    try:
        conn = await _db_pool.get_connection()
        cursor = await conn.execute("SELECT MAX(updated_at) FROM dashboard_stats")
        row = await cursor.fetchone()
        last_stats_ts = row[0] if row and row[0] else 0
    except Exception as e:
        logger.error(f"get_bot_status stats error: {e}")
        last_stats_ts = 0
    bot_alive = (time.time() - last_stats_ts) < 90
    return {
        "frozen": frozen,
        "public_mode": public_mode,
        "last_stats_ts": last_stats_ts,
        "bot_alive": bot_alive,
    }


async def _notify_user(uid: int, text: str):
    """إشعار تيليجرام لمستخدم (بوت مؤقت مثل لوحة التحكم) — نفس رسائل التدفق الحالية"""
    try:
        from telegram import Bot
        bot = Bot(token=config.bot_token)
        await bot.send_message(uid, text, parse_mode="HTML")
    except Exception:
        pass


async def set_bot_frozen_state(frozen: bool, actor_uid: int, actor_name: str) -> Tuple[bool, str]:
    """تجميد / إلغاء تجميد البوت (مالك فقط)"""
    assert_side_effect_safe()
    if actor_uid != config.admin_id:
        return (False, "غير مصرح")
    try:
        await set_bot_frozen(bool(frozen))
        action_type = "BOT_FREEZE" if frozen else "BOT_UNFREEZE"
        await log_admin_action(admin_id=actor_uid, admin_name=actor_name, action_type=action_type,
                               details=f"frozen={frozen}")
        admin_trace(action_type, f"Bot {'frozen' if frozen else 'unfrozen'} by {actor_name}")
        if frozen:
            return (True, "✅ تم تجميد البوت.")
        return (True, "✅ تم إلغاء تجميد البوت.")
    except Exception as e:
        logger.error(f"set_bot_frozen_state error: {e}")
        return (False, f"❌ خطأ: {e}")


async def set_public_mode_state(public: bool, actor_uid: int, actor_name: str) -> Tuple[bool, str]:
    """تفعيل / تعطيل الوضع العام (أدمن أو مالك) — نفس رسالة admin_toggle_mode"""
    assert_side_effect_safe()
    if not await is_admin(actor_uid):
        return (False, "غير مصرح")
    try:
        await set_public_mode(bool(public))
        action_type = "BOT_MODE_PUBLIC" if public else "BOT_MODE_PRIVATE"
        await log_admin_action(admin_id=actor_uid, admin_name=actor_name, action_type=action_type,
                               details=f"public_mode={public}")
        admin_trace(action_type, f"Bot mode set to public={public} by {actor_name}")
        mode = "🌍 عام" if public else "🔐 خاص"
        return (True, f"✅ تم تغيير وضع البوت إلى: {mode}")
    except Exception as e:
        logger.error(f"set_public_mode_state error: {e}")
        return (False, f"❌ خطأ: {e}")


# ==============================================================================
# إدارة الأدمنز (Admin Management) — مشترك مع لوحة التحكم
# ==============================================================================

async def list_admins(actor_uid: int) -> Tuple[bool, List[dict], str]:
    """قائمة الأدمنز (مالك فقط)"""
    if actor_uid != config.admin_id:
        return (False, [], "غير مصرح")
    try:
        admins = await get_all_full_admins()
        result = []
        for (aid, name, tg_username, credit, created_at) in admins:
            u = await db_get_user(aid) or {}
            sub_resellers = await get_admin_sub_resellers(aid)
            result.append({
                "uid": aid,
                "name": name,
                "tg_username": tg_username,
                "is_owner": (aid == config.admin_id),
                "full_admin": u.get("is_admin", 0),
                "sub_resellers_count": len(sub_resellers),
                "credit": credit,
            })
        return (True, result, "")
    except Exception as e:
        logger.error(f"list_admins error: {e}")
        return (False, [], f"❌ خطأ: {e}")


async def add_admin(target_uid: int, actor_uid: int, actor_name: str) -> Tuple[bool, str]:
    """ترقية مستخدم إلى أدمن + اشتراك شهري (مالك فقط) — نفس منطق admin_add_admin_done"""
    assert_side_effect_safe()
    try:
        if actor_uid != config.admin_id:
            return (False, "⛔ هذه الصلاحية للمالك فقط.")

        current_admins = await get_all_full_admins()
        if len(current_admins) >= config.max_full_admins:
            return (False, f"❌ تم الوصول للحد الأقصى من الأدمنز ({config.max_full_admins}).\n"
                           "احذف أدمن أولاً أو عدّل MAX_FULL_ADMINS في .env")

        target_user = await db_get_user(target_uid)
        if not target_user:
            return (False, "❌ المستخدم غير موجود في البوت.")

        if target_uid != config.admin_id and (await is_admin(target_uid) or target_user.get('is_admin', 0) >= 1):
            return (False, "✅ هذا المستخدم أدمن بالفعل.")

        # ✅ تحديث صلاحيات المستخدم + إنشاء اشتراك عادي للأدمن الجديد
        await db_set_user(target_uid, joined_hijri=now_hijri())
        ok = await promote_to_admin(target_uid)
        if not ok:
            return (False, "❌ فشل في الترقية.")

        u = await db_get_user(target_uid) or {}
        cur_exp = u.get("expiry_ts", 0) or 0
        if cur_exp < time.time():
            cur_exp = time.time()
        end_date = cur_exp + (30 * 86400)
        await create_user_subscription(target_uid, "monthly", cur_exp, end_date)

        name = u.get("name", "") or str(target_uid)
        await _notify_user(
            target_uid,
            "👑 <b>تم تعيينك كأدمن في النظام!</b>\n\n"
            "📦 تم تفعيل اشتراك شهري لك (100 واجب).\n"
            "📅 يمكنك تجديد اشتراكك من لوحة التحكم.",
        )

        await log_admin_action(admin_id=actor_uid, admin_name=actor_name, action_type="ADMIN_ADD",
                               target_user_id=target_uid, target_user_name=name,
                               details="Promoted to admin + monthly subscription (30 days)")
        admin_trace("ADMIN_ADD", f"User {target_uid} promoted to admin by {actor_name}", uid=str(target_uid))

        return (True, f"✅ تم ترقية <code>{target_uid}</code> إلى أدمن مع اشتراك شهري.")
    except Exception as e:
        logger.error(f"add_admin error: {e}")
        return (False, f"❌ فشل: {e}")


async def delete_admin(target_uid: int, actor_uid: int, actor_name: str) -> Tuple[bool, str]:
    """حذف أدمن (مالك فقط) — نفس منطق admin_handle_delete_admin"""
    assert_side_effect_safe()
    try:
        if actor_uid != config.admin_id:
            return (False, "❌ هذا الإجراء للمالك فقط.")
        if target_uid == config.admin_id:
            return (False, "❌ لا يمكنك حذف نفسك!")

        target_user = await db_get_user(target_uid)
        if not target_user:
            return (False, "❌ المستخدم غير موجود.")

        if target_user.get('is_admin', 0) < 1 and target_user.get('role') != 'admin':
            return (False, "❌ هذا المستخدم ليس أدمن.")

        ok = await demote_from_admin(target_uid)
        if not ok:
            return (False, "❌ فشل في الحذف.")

        name = target_user.get('real_name') or target_user.get('name') or str(target_uid)
        await _notify_user(
            target_uid,
            "⚠️ <b>تم إزالة صفة الأدمن من حسابك.</b>\n\n"
            "لم تعد تملك لوحة الإدارة أو الصلاحيات الإدارية.",
        )

        await db_log(actor_uid, "DELETE_ADMIN", detail=f"Deleted admin {target_uid}")
        await log_admin_action(admin_id=actor_uid, admin_name=actor_name, action_type="ADMIN_DELETE",
                               target_user_id=target_uid, target_user_name=name,
                               details="Deleted admin")
        admin_trace("ADMIN_DELETE", f"Admin {target_uid} deleted by {actor_name}", uid=str(target_uid))

        return (True, f"✅ <b>تم حذف الأدمن بنجاح!</b>\n\n"
                      f"👤 الأدمن: {name}\n"
                      f"🆔 المعرف: {target_uid}")
    except Exception as e:
        logger.error(f"delete_admin error: {e}")
        return (False, f"❌ فشل: {e}")


async def charge_admin_credit(target_uid: int, amount: float, actor_uid: int, actor_name: str) -> Tuple[bool, str]:
    """شحن رصيد أدمن (مالك فقط) — نفس منطق _admin_charge_admin_user_input + admin_reseller_credit_amount_input"""
    assert_side_effect_safe()
    try:
        if actor_uid != config.admin_id:
            return (False, "❌ هذا الإجراء للمالك فقط.")

        user = await db_get_user(target_uid)
        if not user:
            return (False, "❌ المستخدم غير موجود.")
        if user.get('is_admin', 0) < 1:
            return (False, "❌ هذا المستخدم ليس أدمن.")

        if not isinstance(amount, (int, float)) or isinstance(amount, bool) or not math.isfinite(amount) or amount < 0:
            return (False, "❌ المبلغ غير صالح.")

        ok = await add_reseller_credit(target_uid, amount, details=f"Added by admin {actor_uid}")
        if not ok:
            return (False, "❌ فشل في الشحن.")

        new_balance = await get_reseller_credit(target_uid)
        name = user.get('real_name') or user.get('name') or str(target_uid)
        await _notify_user(
            target_uid,
            f"💰 <b>تم شحن رصيدك!</b>\n\n"
            f"💳 تم إضافة: {amount} credit\n"
            f"💰 رصيدك الحالي: {new_balance} credit",
        )

        await db_log(actor_uid, "ADD_RESELLER_CREDIT", detail=f"Added {amount} to {target_uid}")
        await log_admin_action(admin_id=actor_uid, admin_name=actor_name, action_type="ADMIN_CHARGE",
                               target_user_id=target_uid, target_user_name=name,
                               details=f"Added {amount} credit to admin")
        admin_trace("ADMIN_CHARGE", f"Admin {target_uid} charged +{amount} by {actor_name}", uid=str(target_uid))

        return (True, f"✅ <b>تم الشحن بنجاح!</b>\n\n"
                      f"👤 الموزع: {name}\n"
                      f"💳 تم إضافة: {amount} credit\n"
                      f"💰 الرصيد الجديد: {new_balance} credit")
    except Exception as e:
        logger.error(f"charge_admin_credit error: {e}")
        return (False, f"❌ خطأ: {e}")


# ==============================================================================
# إدارة الموزعين (Reseller Management) — مشترك مع لوحة التحكم
# ==============================================================================

async def list_resellers(actor_uid: int) -> Tuple[bool, List[dict], str]:
    """قائمة الموزعين (أدمن) — عبر db_all_users + فلتر role == 'reseller' + get_reseller_stats"""
    if not await is_admin(actor_uid):
        return (False, [], "غير مصرح")
    try:
        users = await db_all_users()
        result = []
        for u in users:
            uid_i = u.get("telegram_id")
            if u.get("role") != "reseller":
                continue
            stats = await get_reseller_stats(uid_i)
            result.append({
                "uid": uid_i,
                "name": u.get("real_name") or u.get("name") or str(uid_i),
                "credit": u.get("reseller_credit", 0),
                "customers_count": stats.get("total_customers", 0),
                "stats": stats,
            })
        result.sort(key=lambda r: r["credit"], reverse=True)
        return (True, result, "")
    except Exception as e:
        logger.error(f"list_resellers error: {e}")
        return (False, [], f"❌ خطأ: {e}")


async def add_reseller(target_uid: int, actor_uid: int, actor_name: str) -> Tuple[bool, str]:
    """ترقية مستخدم إلى موزع / رئيس موردين (أدمن) — نفس منطق admin_add_reseller_input"""
    assert_side_effect_safe()
    try:
        if not await is_admin(actor_uid):
            return (False, "غير مصرح")

        target_user = await db_get_user(target_uid)
        if not target_user:
            return (False, "❌ المستخدم غير موجود في البوت.")

        if await is_reseller(target_uid) or target_user.get('is_admin', 0) >= 1:
            return (False, "✅ هذا المستخدم موزع/أدمن بالفعل.")

        is_owner = (actor_uid == config.admin_id)

        # Enforce MAX_FULL_ADMINS limit for owner promotions
        if is_owner:
            current_admins = await get_all_full_admins()
            if len(current_admins) >= config.max_full_admins:
                return (False, f"❌ تم الوصول للحد الأقصى من الأدمنز ({config.max_full_admins}).\n"
                               "احذف أدمن أولاً أو عدّل MAX_FULL_ADMINS في .env")

        if is_owner:
            ok = await promote_to_admin(target_uid)
            role_display = "أدمن (رئيس موردين)"
        else:
            ok = await promote_to_reseller(target_uid)
            role_display = "موزع"

        if not ok:
            return (False, "❌ فشل في الترقية.")

        name = target_user.get('real_name') or target_user.get('name') or str(target_uid)
        await _notify_user(
            target_uid,
            "🎉 <b>مبروك! لقد تمت ترقيتك!</b>\n\n"
            "🔑 يمكنك الآن الوصول إلى لوحة الإدارة من القائمة الرئيسية.\n"
            "💳 احصل على رصيد وابدأ في تفعيل الاشتراكات لعملائك!",
        )

        await db_log(actor_uid, "PROMOTE_USER", detail=f"Promoted user {target_uid} to {role_display}")
        await log_admin_action(admin_id=actor_uid, admin_name=actor_name, action_type="RESELLER_ADD",
                               target_user_id=target_uid, target_user_name=name,
                               details=f"Promoted to {role_display}")
        admin_trace("RESELLER_ADD", f"User {target_uid} promoted to {role_display} by {actor_name}", uid=str(target_uid))

        return (True, f"✅ <b>تمت الترقية بنجاح!</b>\n\n"
                      f"👤 المستخدم: {name}\n"
                      f"🆔 المعرف: {target_uid}\n"
                      f"🏷️ الدور: {role_display}")
    except Exception as e:
        logger.error(f"add_reseller error: {e}")
        return (False, f"❌ فشل: {e}")


async def add_reseller_credit_op(target_uid: int, amount: float, actor_uid: int, actor_name: str) -> Tuple[bool, str]:
    """شحن رصيد موزع (أدمن) — نفس منطق admin_reseller_credit_amount_input"""
    assert_side_effect_safe()
    try:
        if not await is_admin(actor_uid):
            return (False, "غير مصرح")
        if not await is_reseller(target_uid):
            return (False, "❌ هذا المستخدم ليس موزعاً.")

        if not isinstance(amount, (int, float)) or isinstance(amount, bool) or not math.isfinite(amount) or amount < 0:
            return (False, "❌ المبلغ غير صالح.")

        ok = await add_reseller_credit(target_uid, amount, details=f"Added by admin {actor_uid}")
        if not ok:
            return (False, "❌ فشل في الشحن.")

        new_balance = await get_reseller_credit(target_uid)
        target_user = await db_get_user(target_uid) or {}
        name = target_user.get('real_name') or target_user.get('name') or str(target_uid)
        await _notify_user(
            target_uid,
            f"💰 <b>تم شحن رصيدك!</b>\n\n"
            f"💳 تم إضافة: {amount} credit\n"
            f"💰 رصيدك الحالي: {new_balance} credit",
        )

        await db_log(actor_uid, "ADD_RESELLER_CREDIT", detail=f"Added {amount} to {target_uid}")
        await log_admin_action(admin_id=actor_uid, admin_name=actor_name, action_type="RESELLER_CREDIT",
                               target_user_id=target_uid, target_user_name=name,
                               details=f"Added {amount} credit")
        admin_trace("RESELLER_CREDIT", f"Reseller {target_uid} +{amount} credit by {actor_name}", uid=str(target_uid))

        return (True, f"✅ <b>تم الشحن بنجاح!</b>\n\n"
                      f"👤 الموزع: {name}\n"
                      f"💳 تم إضافة: {amount} credit\n"
                      f"💰 الرصيد الجديد: {new_balance} credit")
    except Exception as e:
        logger.error(f"add_reseller_credit_op error: {e}")
        return (False, f"❌ فشل: {e}")


_PLAN_NAME_AR = {'weekly': 'أسبوعي', 'monthly': 'شهري', 'semester': 'ترم'}


async def set_reseller_prices_op(prices: dict, actor_uid: int, actor_name: str) -> Tuple[bool, str]:
    """تحديد أسعار Credit (أدمن) — يقبل مفاتيح عربية (أسبوعي/شهري/ترم) أو إنجليزية — نفس منطق admin_reseller_prices_input"""
    assert_side_effect_safe()
    try:
        if not await is_admin(actor_uid):
            return (False, "غير مصرح")

        plan_map = {
            'أسبوعي': 'weekly',
            'شهري': 'monthly',
            'ترم': 'semester',
        }

        # تطبيع المفاتيح (عربية أو إنجليزية) + تحقق من القيم
        normalized = []
        for key, val in prices.items():
            plan_type = plan_map.get(key, key)
            if plan_type not in _PLAN_NAME_AR:
                return (False, "❌ اسم الخطة غير صحيح. استخدم: أسبوعي، شهري، ترم")
            if not isinstance(val, (int, float)) or isinstance(val, bool) or not math.isfinite(val) or val <= 0:
                return (False, "❌ السعر يجب أن يكون أكبر من 0.")
            normalized.append((plan_type, val))

        for plan_type, price in normalized:
            ok = await set_reseller_credit_price(plan_type, price)
            if not ok:
                return (False, "❌ فشل في تغيير السعر.")

        lines = []
        for plan_type, price in normalized:
            lines.append(f"📦 الخطة: {_PLAN_NAME_AR[plan_type]}\n💰 السعر الجديد: {price} credit")
            await db_log(actor_uid, "SET_RESELLER_PRICE", detail=f"{plan_type}: {price}")
        details = ", ".join(f"{plan_type}: {price}" for plan_type, price in normalized)
        await log_admin_action(admin_id=actor_uid, admin_name=actor_name, action_type="RESELLER_PRICES",
                               details=details)
        admin_trace("RESELLER_PRICES", f"Prices set ({details}) by {actor_name}")

        return (True, "✅ <b>تم تغيير السعر!</b>\n\n" + "\n\n".join(lines))
    except Exception as e:
        logger.error(f"set_reseller_prices_op error: {e}")
        return (False, f"❌ فشل: {e}")


async def reseller_stats_op(actor_uid: int) -> Tuple[bool, dict, str]:
    """إحصائيات الموزعين الإجمالية (أدمن) — نفس تجميع admin_reseller_stats_panel"""
    if not await is_admin(actor_uid):
        return (False, {}, "غير مصرح")
    try:
        conn = await _db_pool.get_connection()

        cursor = await conn.execute("SELECT COUNT(*) FROM users WHERE role = 'reseller'")
        total = (await cursor.fetchone())[0]

        cursor = await conn.execute(
            "SELECT COALESCE(SUM(reseller_credit), 0) FROM users WHERE role = 'reseller'"
        )
        total_credit = (await cursor.fetchone())[0]

        cursor = await conn.execute("SELECT COUNT(*) FROM reseller_keys")
        total_keys = (await cursor.fetchone())[0]

        cursor = await conn.execute("SELECT COUNT(*) FROM reseller_keys WHERE used = 1")
        total_used = (await cursor.fetchone())[0]

        cursor = await conn.execute(
            "SELECT COUNT(*) FROM users WHERE referred_by_reseller IS NOT NULL"
        )
        total_customers = (await cursor.fetchone())[0]

        prices = await get_all_reseller_credit_prices()

        return (True, {
            "total": total,
            "total_credit": total_credit,
            "total_keys": total_keys,
            "total_used": total_used,
            "total_customers": total_customers,
            "weekly": prices.get("weekly", 0),
            "monthly": prices.get("monthly", 0),
            "semester": prices.get("semester", 0),
        }, "")
    except Exception as e:
        logger.error(f"reseller_stats_op error: {e}")
        return (False, {}, f"❌ خطأ: {e}")


async def delete_reseller_op(target_uid: int, actor_uid: int, actor_name: str) -> Tuple[bool, str]:
    """حذف موزع (مالك فقط) — نفس منطق admin_handle_delete_reseller"""
    assert_side_effect_safe()
    try:
        if actor_uid != config.admin_id:
            return (False, "❌ هذا الإجراء للمالك فقط.")
        if target_uid == config.admin_id:
            return (False, "❌ لا يمكنك حذف نفسك!")

        target_user = await db_get_user(target_uid)
        if not target_user:
            return (False, "❌ المستخدم غير موجود.")

        if not await is_reseller(target_uid):
            return (False, "❌ هذا المستخدم ليس موزعاً.")

        ok = await demote_from_reseller(target_uid)
        if not ok:
            return (False, "❌ فشل في الحذف.")

        # دفاعي: لو كان الموزع رئيس موردين (is_admin) تُزال صلاحية الأدمن أيضاً لضمان حذف كل الصلاحيات
        if target_user.get('is_admin', 0) >= 1:
            await demote_from_admin(target_uid)

        name = target_user.get('real_name') or target_user.get('name') or str(target_uid)
        await _notify_user(
            target_uid,
            "⚠️ <b>تم إزالة صفة الموزع من حسابك.</b>\n\n"
            "لم تعد تملك لوحة الموزع أو الرصيد.",
        )

        await db_log(actor_uid, "DELETE_RESELLER", detail=f"Deleted reseller {target_uid}")
        await log_admin_action(admin_id=actor_uid, admin_name=actor_name, action_type="RESELLER_DELETE",
                               target_user_id=target_uid, target_user_name=name,
                               details="Deleted reseller")
        admin_trace("RESELLER_DELETE", f"Reseller {target_uid} deleted by {actor_name}", uid=str(target_uid))

        return (True, f"✅ <b>تم حذف الموزع بنجاح!</b>\n\n"
                      f"👤 الموزع: {name}\n"
                      f"🆔 المعرف: {target_uid}")
    except Exception as e:
        logger.error(f"delete_reseller_op error: {e}")
        return (False, f"❌ فشل: {e}")


async def ban_reseller_customer_op(customer_uid: int, action: str, actor_uid: int, actor_name: str) -> Tuple[bool, str]:
    """حظر / إيقاف اشتراك عميل موزع (مالك فقط) — نفس منطق admin_handle_ban_reseller_customer_callback"""
    assert_side_effect_safe()
    try:
        if actor_uid != config.admin_id:
            return (False, "❌ هذا الإجراء للمالك فقط.")

        if action == 'cancel':
            return (False, "❌ تم الإلغاء.")
        if action not in ('ban', 'stop'):
            return (False, "❌ إجراء غير صالح.")

        target_user = await db_get_user(customer_uid)
        if not target_user:
            return (False, "❌ المستخدم غير موجود.")

        customer_name = target_user.get('real_name') or target_user.get('name') or str(customer_uid)

        conn = await _db_pool.get_connection()

        if action == 'ban':
            # Ban user: set free_attempts to 0 and deactivate subscription
            await conn.execute(
                "UPDATE users SET free_attempts = 0, expiry_ts = 0 WHERE telegram_id = ?",
                (customer_uid,)
            )
            await conn.execute(
                "UPDATE user_subscriptions SET is_active = 0 WHERE user_id = ? AND is_active = 1",
                (customer_uid,)
            )
            await conn.commit()

            await _notify_user(
                customer_uid,
                "🚫 <b>تم إيقاف حسابك</b>\n\n"
                "لقد تم إلغاء اشتراكك وتصفير رصيدك.\n"
                "تواصل مع الإدارة لمعرفة السبب.",
            )

            msg = (f"🚫 <b>تم حظر المستخدم!</b>\n\n"
                   f"👤 العميل: {customer_name}\n"
                   f"✅ تم إيقاف الاشتراك وتصفير الرصيد")
        else:  # stop
            # Stop subscription only
            await conn.execute(
                "UPDATE user_subscriptions SET is_active = 0 WHERE user_id = ? AND is_active = 1",
                (customer_uid,)
            )
            await conn.execute(
                "UPDATE users SET expiry_ts = 0 WHERE telegram_id = ?",
                (customer_uid,)
            )
            await conn.commit()

            await _notify_user(
                customer_uid,
                "⏹️ <b>تم إيقاف اشتراكك</b>\n\n"
                "لم يعد لديك اشتراك نشط.\n"
                "تواصل مع الإدارة لمعرفة التفاصيل.",
            )

            msg = (f"⏹️ <b>تم إيقاف الاشتراك!</b>\n\n"
                   f"👤 العميل: {customer_name}\n"
                   f"✅ تم إلغاء الاشتراك النشط")

        await db_log(actor_uid, "BAN_RESELLER_CUSTOMER",
                     detail=f"Action: {action}, Target: {customer_uid}")
        await log_admin_action(admin_id=actor_uid, admin_name=actor_name, action_type="RESELLER_BAN",
                               target_user_id=customer_uid, target_user_name=customer_name,
                               details=f"Action: {action}")
        admin_trace("RESELLER_BAN", f"Customer {customer_uid} {action} by {actor_name}", uid=str(customer_uid))

        return (True, msg)
    except Exception as e:
        logger.error(f"ban_reseller_customer_op error: {e}")
        return (False, f"❌ خطأ: {e}")
