#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
HASAD Bot - Main entry point
"""
import asyncio
import sys
import warnings
import csv
from hasad_bot.datetime_utils import now, now_timestamp, format_datetime, datetime, timedelta
from pathlib import Path
import datetime as dt
from hasad_bot.logger import log_function_call
import time
import traceback
from hasad_bot.handlers import cmd_start_tunnel, cmd_stop_tunnel, cmd_tunnel_status, cmd_dashboard_url, cmd_user_log, cmd_announce, cb_link_help, cb_link_nudge_back
from hasad_bot import datetime_utils
from hasad_bot.utils import now_hijri
from hasad_bot.admin_ops import (send_encrypted_excel_file, send_encrypted_zip_file, send_encrypted_file, send_db_backup, send_cv_export, extract_credentials_terminal)
# أضف المسار الحالي
sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from telegram import Update
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ConversationHandler,
    PreCheckoutQueryHandler,
    ContextTypes,
    filters,
)
from telegram.warnings import PTBUserWarning
from loguru import logger




from hasad_bot.handlers import (
    # الدوال الأساسية (الموجودة)
    start,
    handle_text,
    _cancel_handler,
    error_handler,
    login_got_username,
    login_got_password,
    admin_renew_got_user,
    admin_renew_got_days,
    admin_revoke_done,
    admin_genkeys_done,
    admin_add_admin_done,
    admin_send_reply_done,
    admin_broadcast_send,
    cmd_admin_panel,
    handle_admin_password,
    support_msg_handler,
    exit_support_room,
    engine_callback_handler,
    cb_user_detail,
    cb_unlock,
    cb_delete,
    cb_back,
    cb_request_unlock,
    cb_view_support_history,
    cb_reply_support,
    solve_homework,
    share_and_earn,
    my_account,
    cmd_login,
    activate_subscription,
    enter_support_room,
    admin_panel,
    admin_system_stats,
    admin_extract_credentials,
    admin_broadcast_ask,
    admin_renew_ask,
    admin_revoke_ask,
    admin_genkeys_ask,
    admin_toggle_mode,
    admin_list_users,
    admin_add_admin_ask,
    admin_files,
    admin_full_reset,
    help_command,
    broadcast_target_callback,
    
    # دوال فك القفل
    cb_unlock_approve,
    cb_unlock_reject,
    cb_unlock_reason,
    cb_unlock_back,
    cb_unlock_cancel,
    cb_unlock_custom_reason,
    
    # ✅ الدوال الجديدة للبث الجماعي

    
    # ✅ دوال إضافة واجبات (جديدة)
    admin_add_homework_start,
    admin_add_hw_got_id,
    admin_add_hw_got_count,
    admin_add_hw_confirm_callback,
    
    # دوال المتجر والدفع (موجودة)
    open_shop,
    shop_plan_callback,
    shop_pay_callback,
    shop_back_callback,
    successful_payment_handler,
    pre_checkout_handler,
    select_school_callback,
    webapp_data_handler,
    admin_panel_web,
    admin_settings,
    activate_request_callback,
    set_days_callback,
    reject_request_callback,
    reject_reason_callback,
    show_all_requests_callback,
    view_request_callback,
    back_to_request_callback,
    back_to_main_callback,
    custom_days_callback,
    handle_custom_days_input,
    handle_custom_days,
    handle_custom_reason,
    reject_custom_callback,
    handle_custom_reject,
    back_to_account_callback,
    view_day_report_callback,
    show_reports_list_callback,
    get_all_payment_requests,
)

from hasad_bot.handlers.reseller import (
    reseller_panel,
    reseller_customers,
    reseller_activate,
    reseller_stats,
    reseller_link,
    reseller_activate_callback,
    reseller_select_customer_callback,
    reseller_tx_log,
)


from hasad_bot.handlers.admin_reseller import (
    admin_reseller_panel,
    admin_add_reseller,
    admin_add_reseller_input,
    admin_reseller_credit,
    admin_reseller_credit_user_input,
    admin_reseller_credit_amount_input,
    admin_reseller_list,
    admin_reseller_prices,
    admin_reseller_prices_input,
    admin_reseller_stats_panel,
    admin_delete_reseller,
    admin_ban_reseller_customer,
    admin_handle_delete_reseller,
    admin_handle_ban_reseller_customer,
    admin_handle_ban_reseller_customer_callback,
)














# ==============================================================================
# استيرادات من hasad_bot
# ==============================================================================

from hasad_bot.config import (
    config,
    MAIN_MENU,
    ADMIN_PANEL,
    AWAIT_LOGIN_USERNAME,
    AWAIT_LOGIN_PASSWORD,
    AWAIT_RENEW_USER,
    AWAIT_RENEW_DAYS,
    AWAIT_REVOKE_USER,
    AWAIT_GENKEY_COUNT,
    AWAIT_ADD_ADMIN,
    AWAIT_SUPPORT_MSG,
    AWAIT_ADMIN_REPLY,
    AWAIT_BROADCAST_MSG,
    AWAIT_BROADCAST_CONFIRM,
    AWAIT_ADD_HW_ID,
    AWAIT_ADD_HW_COUNT,
    AWAIT_ADD_HW_CONFIRM,
    AWAIT_CUSTOM_DAYS,
    AWAIT_BROADCAST_TARGET,
    AWAIT_CUSTOM_REASON,
    AWAIT_RESELLER_CREDIT_AMOUNT,
    AWAIT_RESELLER_CREDIT_USER,
    AWAIT_RESELLER_ACTIVATE_USER,
    AWAIT_RESELLER_PRICES,
    AWAIT_ADMIN_PASSWORD,
)

from hasad_bot.database import (
    db_init,
    db_set_user,
    db_get_user,
    db_all_users,
    _db_pool
)


from hasad_bot.playwright_engine import _browser_pool
from hasad_bot.radar_engine import handle_radar_callback
from hasad_bot.utils import BANNER_TERMINAL, admin_trace, now_hijri, decrypt_password, encrypt_password
# main.py - استبدل السطر 167 بهذا

# ==============================================================================
# حفظ اللوج المباشر في ملف
# ==============================================================================



# ==============================================================================
# دوال البوت الأساسية
# ==============================================================================














async def post_init(application: Application):
    """Post initialization hook - النسخة المتكاملة مع جميع التحسينات"""
    
    logger.info("🚀 Initializing HASAD Bot...")
    



    from hasad_bot.database import collect_and_save_dashboard_stats

    async def update_dashboard_stats_job(context: ContextTypes.DEFAULT_TYPE):
        """تحديث إحصائيات الداشبورد كل 15 ثواني"""
        await collect_and_save_dashboard_stats()

# بعد تهيئة كل شيء
    job_queue = application.job_queue
    if job_queue:
    # run_repeating(الدالة, الفاصل_بالثواني, first=التأخير_قبل_أول_تشغيل)
        job_queue.run_repeating(update_dashboard_stats_job, interval=15, first=3)
        logger.info("📊 Dashboard stats job started (every 15 seconds)")

        # ============================================================
        # 📢 إعلانات متكررة ومستهدفة (Marketing Automation)
        # ============================================================
        from datetime import time as dtime
        from hasad_bot.handlers import (
            ensure_announcement_tables,
            send_announcement,
            AnnouncementType,
        )

        # تهيئة الجداول والقوالب
        await ensure_announcement_tables()
        logger.info("📢 Announcement tables ensured")

        # جدولة الإعلانات اليومية
        async def _scheduled_free_promo(context):
            try:
                await send_announcement(context.bot, AnnouncementType.FREE_USER_PROMO.value, manual=False)
            except Exception as e:
                logger.error(f"Free promo job failed: {e}")

        async def _scheduled_expiring_5d(context):
            try:
                await send_announcement(context.bot, AnnouncementType.SUB_EXPIRING_5D.value, manual=False)
            except Exception as e:
                logger.error(f"Expiring 5d job failed: {e}")

        async def _scheduled_expiring_1d(context):
            try:
                await send_announcement(context.bot, AnnouncementType.SUB_EXPIRING_1D.value, manual=False)
            except Exception as e:
                logger.error(f"Expiring 1d job failed: {e}")

        async def _scheduled_low_attempts(context):
            try:
                await send_announcement(context.bot, AnnouncementType.LOW_ATTEMPTS.value, manual=False)
            except Exception as e:
                logger.error(f"Low attempts job failed: {e}")

        async def _scheduled_share_earn(context):
            try:
                await send_announcement(context.bot, AnnouncementType.SHARE_AND_EARN_PROMO.value, manual=False)
            except Exception as e:
                logger.error(f"Share & earn job failed: {e}")

        async def _scheduled_link_reminder(context):
            try:
                await send_announcement(context.bot, AnnouncementType.LINK_REMINDER.value, manual=False)
            except Exception as e:
                logger.error(f"Link reminder job failed: {e}")

        # run_daily(time, days) — أيام 0=الإثنين ... 6=الأحد
        job_queue.run_daily(_scheduled_free_promo, time=dtime(10, 0), name="announce_free_promo")
        job_queue.run_daily(_scheduled_expiring_5d, time=dtime(18, 0), name="announce_expiring_5d")
        job_queue.run_daily(_scheduled_expiring_1d, time=dtime(19, 0), name="announce_expiring_1d")
        job_queue.run_daily(_scheduled_low_attempts, time=dtime(9, 0), name="announce_low_attempts")
        job_queue.run_daily(_scheduled_share_earn, time=dtime(14, 0), name="announce_share_earn")
        job_queue.run_daily(_scheduled_link_reminder, time=dtime(12, 0), name="announce_link_reminder")
        logger.info("📢 Daily announcement jobs scheduled (6 types)")
    
    # ============================================================
    # 1️⃣ تهيئة قاعدة البيانات
    # ============================================================
    await db_init()
    logger.info("✅ Database initialized")
    
    # ============================================================
    # 2️⃣ تهيئة الـ Logger
    # ============================================================
    from hasad_bot.logger import init_logger
    await init_logger(_db_pool)
    logger.info("✅ Logger initialized")
    
    # ============================================================
    # 3️⃣ تهيئة المتصفح (Browser Pool)
    # ============================================================
    from hasad_bot.playwright_engine import _browser_pool
    logger.info("🚀 Initializing browser pool...")
    await _browser_pool.initialize()
    logger.info("✅ Browser pool initialized")
    
    # ============================================================
    # 4️⃣ تشغيل الرادار (مُعطّل مؤقتاً بطلب من المستخدم)
    # ============================================================
    if config.radar_enabled:
        from hasad_bot.radar_engine import radar_engine
        await radar_engine.start(application.bot)
        logger.info("✅ Radar engine started")
    else:
        logger.warning("⏸ Radar engine disabled (RADAR_ENABLED=false)")
    
    # ============================================================
    # 5️⃣ تصفير الحالات المعلقة (State Clear)
    # ============================================================
    try:
        from hasad_bot.ai_engine import active_sessions
        active_sessions.clear()
        logger.info("✅ تم تصفير جميع جلسات المستخدمين المعلقة")
    except Exception as e:
        logger.warning(f"⚠️ فشل تصفير الجلسات: {e}")
    
    # ============================================================
    # 6️⃣ إنشاء حساب المالك إذا لم يكن موجوداً
    # ============================================================
    u = await db_get_user(config.admin_id)
    if not u:
        await db_set_user(
            config.admin_id,
            name="Admin",
            tg_username="admin",
            dars360_user=config.admin_dars_user,
            dars360_pass=encrypt_password(config.admin_dars_pass),
            expiry_ts=9_999_999_999.0,
            expiry_hijri="دائم ♾️",
            locked_to=config.admin_id,
            is_admin=2,
            joined_hijri=now_hijri(),
            rank_title="👑 المالك"
        )
        
        from hasad_bot.database import create_user_subscription
        start_date = time.time()
        end_date = start_date + (120 * 86400)
        
        conn = await _db_pool.get_connection()
        await conn.execute("""
            INSERT INTO user_subscriptions 
            (user_id, plan_id, start_date, end_date, max_homeworks, homeworks_used, is_active)
            VALUES (?, 'semester', ?, ?, 200, 0, 1)
        """, (config.admin_id, start_date, end_date))
        await conn.commit()
        
        logger.info(f"✅ Super admin created: {config.admin_id}")
    
    # ============================================================
    # 7️⃣ إضافة الأعمدة الجديدة للقاعدة القديمة (platform_url, platform_id)
    # ============================================================
    try:
        conn = await _db_pool.get_connection()
        cursor = await conn.execute("PRAGMA table_info(users)")
        columns = await cursor.fetchall()
        column_names = [col[1] for col in columns]
        
        if 'platform_url' not in column_names:
            await conn.execute("ALTER TABLE users ADD COLUMN platform_url TEXT DEFAULT ''")
            logger.info("✅ تم إضافة عمود: platform_url")
        
        if 'platform_id' not in column_names:
            await conn.execute("ALTER TABLE users ADD COLUMN platform_id TEXT DEFAULT ''")
            logger.info("✅ تم إضافة عمود: platform_id")
        
        await conn.commit()
        
        # تحديث المستخدمين القدامى بالمدرسة الافتراضية
        await conn.execute("""
            UPDATE users 
            SET platform_url = 'https://alamjad1.dars360.com',
                platform_id = 'alamjad1'
            WHERE dars360_user IS NOT NULL 
            AND dars360_user != ''
            AND (platform_url IS NULL OR platform_url = '')
        """)
        await conn.commit()
        
        cursor = await conn.execute("SELECT changes()")
        count = await cursor.fetchone()
        if count and count[0] > 0:
            logger.info(f"✅ تم تحديث {count[0]} مستخدم قديم بالمدرسة الافتراضية")
        
    except Exception as e:
        logger.warning(f"⚠️ لم نتمكن من إضافة الأعمدة: {e}")
    
    # ============================================================
    # 8️⃣ تسجيل وقت العودة (Uptime Log)
    # ============================================================
    import time as time_module
    from datetime import datetime
    
    uptime_log = f"""
{'='*60}
✅ SYSTEM IS BACK ONLINE
📅 التاريخ الميلادي: {now().strftime('%Y-%m-%d %H:%M:%S')}
📅 التاريخ الهجري: {now_hijri()}
⏱️ وقت التشغيل: {time_module.strftime('%H:%M:%S')}
{'='*60}
"""
    
    logger.success(f"✅ System is back online after maintenance")
    
    # إرسال إشعار للإدارة (اختياري)
    try:
        await application.bot.send_message(
            chat_id=config.admin_id,
            text=f"✅ **البوت رجع شغال**\n📅 {now_hijri()}\n⏱️ {now().strftime('%H:%M:%S')}",
            parse_mode="Markdown"
        )
    except Exception as e:
        logger.warning(f"⚠️ فشل إرسال إشعار العودة: {e}")
    
    # ============================================================
    # 9️⃣ مراقبة الأداء (Performance Monitoring)
    # ============================================================
    try:
        import psutil
        
        cpu = psutil.cpu_percent(interval=1)
        memory = psutil.virtual_memory().percent
        
        logger.info(f"📊 System Performance: CPU={cpu}%, RAM={memory}%")
        
        # إذا كان الأداء عالياً، أرسل تنبيه للإدارة
        if cpu > 80 or memory > 80:
            await application.bot.send_message(
                chat_id=config.admin_id,
                text=f"⚠️ **تنبيه أداء عالي**\n⚙️ CPU: {cpu}%\n💾 RAM: {memory}%",
                parse_mode="Markdown"
            )
            
    except ImportError:
        logger.warning("⚠️ psutil not installed, performance monitoring disabled")
    except Exception as e:
        logger.warning(f"⚠️ Performance monitoring failed: {e}")
    
    # ============================================================
    # 🔟 تشغيل لوحة التحكم (Web Dashboard) في الخلفية
    # ============================================================
    try:
        import subprocess
        import sys
        from pathlib import Path

        dashboard_script = Path(__file__).parent / "hasad_bot" / "web_dashboard.py"

        if dashboard_script.exists():
            # ✅ dashboard.py يبحث عن منفذ متاح داخلياً (find_working_port)
            #    فيتعامل مع المنافذ المحجوزة على Windows
            subprocess.Popen(
                [sys.executable, str(dashboard_script)],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0
            )
            logger.info("✅ Web Dashboard started in background")
            print(f"\n🌐 لوحة التحكم: http://127.0.0.1:{config.dashboard_port} (أو منفذ بديل لو محجوز)")
        else:
            logger.warning(f"⚠️ Web Dashboard file not found: {dashboard_script}")
            
    except Exception as e:
        logger.error(f"Failed to start dashboard: {e}")
    
    # ============================================================
    # ✅ النهاية
    # ============================================================
    logger.success("✅ HASAD Bot initialized successfully")

    
async def shutdown(application: Application):
    """Clean shutdown"""
    logger.info("🛑 Shutting down HASAD Bot...")

    # Stop radar engine (إن كان مُفعّلاً)
    try:
        if config.radar_enabled:
            from hasad_bot.radar_engine import radar_engine
            await radar_engine.stop()
    except Exception as e:
        logger.warning(f"⚠️ Radar stop warning: {e}")

    await _browser_pool.close()
    logger.success("✅ Shutdown complete")








# ==============================================================================
# TERMINAL COMMAND CENTER
# ==============================================================================

async def terminal_input_listener(bot):
    """
    🎮 TERMINAL COMMAND CENTER - الإصدار الأسطوري
    """
    import sys
    import os
    import asyncio
    import subprocess
    import webbrowser
    import psutil
    import json
    from pathlib import Path
    
    # ====================== الاستيرادات الصحيحة ======================
    from hasad_bot.datetime_utils import now          # ← هذا السطر المهم جداً
    from hasad_bot.config import config
    
    # إذا ما عندك كلاس Colors في utils.py، استخدم هذا
    class Colors:
        HEADER = '\033[95m'
        BLUE = '\033[94m'
        CYAN = '\033[96m'
        GREEN = '\033[92m'
        YELLOW = '\033[93m'
        RED = '\033[91m'
        MAGENTA = '\033[35m'
        WHITE = '\033[97m'
        BOLD = '\033[1m'
        UNDERLINE = '\033[4m'
        END = '\033[0m'
        CLEAR = '\033[2J\033[H'

    # ====================== طباعة البداية ======================
    print(f"{Colors.CLEAR}{Colors.CYAN}{Colors.BOLD}")
    print("╔═══════════════════════════════════════════════════════════════════╗")
    print("║                    🎮 TERMINAL COMMAND CENTER                     ║")
    print("║                      الإصدار الأسطوري V2.0                       ║")
    print("╚═══════════════════════════════════════════════════════════════════╝")
    print(f"{Colors.END}")
    
    print(f"{Colors.YELLOW}📅 {now().strftime('%Y-%m-%d %H:%M:%S')}{Colors.END}")
    print(f"{Colors.GREEN}🚀 النظام شغال وعليه البركة{Colors.END}")
    print("\n" + "="*70)
    print(f"{Colors.BOLD}{Colors.WHITE}📟 الأوامر المتاحة (اكتب HELP للقائمة الكاملة):{Colors.END}")
    print("="*70)

    # باقي الكود يبقى **كما هو عندك** (من commands_preview إلى نهاية الدالة)
    # لا تغير أي شيء بعد هذا السطر    
    commands_preview = [
        ("DB", "نسخة احتياطية"),
        ("CV", "تصدير بيانات الطلاب"),
        ("STATS", "إحصائيات"),
        ("POOL", "المتصفحات"),
        ("WEB", "رابط الداشبورد"),
        ("TUNNEL", "Cloudflare Tunnel"),
        ("HELP", "كل الأوامر")
    ]
    
    preview_line = "   ".join([f"{Colors.GREEN}{cmd}{Colors.END}:{desc}" for cmd, desc in commands_preview])
    print(f"   {preview_line}")
    print("="*70 + "\n")
    
    start_time = now()
    command_history = []
    
    while True:
        try:
            cmd = await asyncio.to_thread(sys.stdin.readline)
            cmd = cmd.strip().upper()
            
            if not cmd:
                continue
            
            command_history.append({'cmd': cmd, 'time': now().strftime('%H:%M:%S')})
            if len(command_history) > 10:
                command_history.pop(0)
            
            if cmd == 'DB':
                logger.info("📦 Creating database backup...")
                print(f"{Colors.YELLOW}⏳ جاري إنشاء النسخة الاحتياطية...{Colors.END}")
                try:
                    await send_db_backup(bot)
                    print(f"{Colors.GREEN}✅ تم إنشاء النسخة الاحتياطية بنجاح{Colors.END}")
                except Exception as e:
                    print(f"{Colors.RED}❌ فشل النسخ الاحتياطي: {e}{Colors.END}")
            
            elif cmd == 'CV':
                logger.info("🕵️ Exporting CV data...")
                print(f"{Colors.YELLOW}⏳ جاري تجهيز ملف CV...{Colors.END}")
                try:
                    await send_cv_export(bot)
                    print(f"{Colors.GREEN}✅ تم تصدير بيانات الطلاب{Colors.END}")
                except Exception as e:
                    print(f"{Colors.RED}❌ فشل التصدير: {e}{Colors.END}")
            
            elif cmd == 'WEB':
                logger.info("🌐 Web dashboard command")
                try:
                    import socket
                    from pathlib import Path
                    from hasad_bot.web_dashboard import find_working_port
                    from hasad_bot.handlers.tunnel import tunnel_manager

                    # ─────────────────────────────────────────────────────────
                    # 1) كشف حالة الـ dashboard المحلي + تشغيله إذا متوقف
                    # ─────────────────────────────────────────────────────────
                    candidate_ports = [config.dashboard_port, 8765, 9876, 9999, 15000]
                    actual_port = None
                    dashboard_running = False

                    for port in candidate_ports:
                        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                        sock.settimeout(0.5)
                        try:
                            if sock.connect_ex(('127.0.0.1', port)) == 0:
                                actual_port = port
                                dashboard_running = True
                                break
                        finally:
                            sock.close()

                    if not dashboard_running:
                        dashboard_script = Path(__file__).parent / "hasad_bot" / "web_dashboard.py"
                        if dashboard_script.exists():
                            print(f"{Colors.YELLOW}⏳ تشغيل الداشبورد المحلي...{Colors.END}")
                            subprocess.Popen(
                                [sys.executable, str(dashboard_script)],
                                stdout=subprocess.DEVNULL,
                                stderr=subprocess.DEVNULL,
                                creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0
                            )
                            await asyncio.sleep(1.5)
                            actual_port = find_working_port(config.dashboard_port, fallbacks=[8765, 9876, 9999, 15000])
                            dashboard_running = (actual_port != config.dashboard_port) or True

                    local_url = f"http://127.0.0.1:{actual_port}" if actual_port else f"http://127.0.0.1:{config.dashboard_port}"

                    # ─────────────────────────────────────────────────────────
                    # 2) كشف حالة الـ Cloudflare Tunnel
                    # ─────────────────────────────────────────────────────────
                    tunnel_status = tunnel_manager.get_status()
                    tunnel_url = tunnel_status.get('tunnel_url')
                    tunnel_running = tunnel_status.get('is_running', False)

                    # ─────────────────────────────────────────────────────────
                    # 3) عرض النتيجة
                    # ─────────────────────────────────────────────────────────
                    print(f"\n{Colors.CYAN}{Colors.BOLD}{'='*70}{Colors.END}")
                    print(f"{Colors.CYAN}{Colors.BOLD}🌐  HASAD Dashboard URLs{Colors.END}")
                    print(f"{Colors.CYAN}{Colors.BOLD}{'='*70}{Colors.END}\n")

                    # المحلي
                    print(f"{Colors.GREEN}📍 الرابط المحلي (Local):{Colors.END}")
                    print(f"   {Colors.BOLD}{Colors.WHITE}{local_url}{Colors.END}")
                    if actual_port and actual_port != config.dashboard_port:
                        print(f"   {Colors.YELLOW}ℹ️  (المنفذ المفضّل {config.dashboard_port} محجوز — استُخدم {actual_port}){Colors.END}")
                    elif not dashboard_running:
                        print(f"   {Colors.RED}❌ الداشبورد لم يبدأ — تحقق من الأخطاء{Colors.END}")

                    # الخارجي
                    print(f"\n{Colors.GREEN}🌍 الرابط الخارجي (Cloudflare Tunnel):{Colors.END}")
                    if tunnel_running and tunnel_url:
                        print(f"   {Colors.BOLD}{Colors.WHITE}{tunnel_url}{Colors.END}")
                        print(f"   {Colors.GREEN}✅ الحالة: شغال{Colors.END}")
                    else:
                        print(f"   {Colors.YELLOW}⏸️  الحالة: متوقف{Colors.END}")
                        # فحص هل في cloudflared مستقل شغّال
                        try:
                            import shutil
                            if shutil.which("cloudflared"):
                                print(f"   💡 للبدء: {Colors.CYAN}TUNNEL START{Colors.END} أو اضغط /start_tunnel في تيليجرام")
                        except Exception:
                            pass

                    print(f"\n{Colors.CYAN}{'='*70}{Colors.END}")

                    # ─────────────────────────────────────────────────────────
                    # 4) فتح الرابط المحلي تلقائياً
                    # ─────────────────────────────────────────────────────────
                    if actual_port and dashboard_running:
                        webbrowser.open(local_url)
                        print(f"{Colors.GREEN}✅ تم فتح الرابط المحلي في المتصفح{Colors.END}")
                    elif actual_port:
                        print(f"{Colors.YELLOW}⚠️ حاول فتح {local_url} يدوياً{Colors.END}")

                    # اختصارات
                    print(f"\n{Colors.CYAN}💡 اختصارات:{Colors.END}")
                    print(f"   • {Colors.GREEN}TUNNEL START{Colors.END}  — بدء Cloudflare Tunnel")
                    print(f"   • {Colors.GREEN}TUNNEL STOP{Colors.END}   — إيقاف Tunnel")
                    print(f"   • {Colors.GREEN}TUNNEL{Colors.END}         — عرض الحالة فقط")
                    print()
                except Exception as e:
                    print(f"{Colors.RED}❌ خطأ في فتح الداشبورد: {e}{Colors.END}")
                    import traceback
                    traceback.print_exc()

            elif cmd == 'TUNNEL':
                # ─────────────────────────────────────────────────────────
                # أوامر Tunnel المختصرة
                # ─────────────────────────────────────────────────────────
                parts = cmd.split()
                action = parts[1].upper() if len(parts) > 1 else 'STATUS'
                from hasad_bot.handlers.tunnel import tunnel_manager

                if action == 'START':
                    print(f"{Colors.YELLOW}🚀 جاري تشغيل Cloudflare Tunnel...{Colors.END}")
                    try:
                        from hasad_bot.handlers.tunnel import tunnel_manager as tm
                        success, result = await tm.start_tunnel(bot, config.admin_id)
                        if success:
                            print(f"{Colors.GREEN}✅ Tunnel شغّال: {result}{Colors.END}")
                        else:
                            print(f"{Colors.RED}❌ {result}{Colors.END}")
                    except Exception as e:
                        print(f"{Colors.RED}❌ خطأ: {e}{Colors.END}")

                elif action == 'STOP':
                    print(f"{Colors.YELLOW}🛑 جاري إيقاف Tunnel...{Colors.END}")
                    try:
                        from hasad_bot.handlers.tunnel import tunnel_manager as tm
                        success, result = await tm.stop_tunnel(bot, config.admin_id)
                        print(f"{Colors.GREEN if success else Colors.RED}{result}{Colors.END}")
                    except Exception as e:
                        print(f"{Colors.RED}❌ خطأ: {e}{Colors.END}")

                else:  # STATUS
                    status = tunnel_manager.get_status()
                    if status['is_running'] and status['tunnel_url']:
                        print(f"{Colors.GREEN}✅ Tunnel شغّال:{Colors.END}")
                        print(f"   {Colors.BOLD}{Colors.WHITE}{status['tunnel_url']}{Colors.END}")
                    else:
                        print(f"{Colors.YELLOW}⏸️  Tunnel متوقف{Colors.END}")
                        print(f"   للبدء: {Colors.CYAN}TUNNEL START{Colors.END}")
            
            elif cmd == 'EX':
                logger.info("🔑 Extracting platform credentials...")
                print(f"{Colors.YELLOW}⏳ جاري استخراج البيانات...{Colors.END}")
                try:
                    await extract_credentials_terminal(bot)
                    print(f"{Colors.GREEN}✅ تم استخراج البيانات{Colors.END}")
                except Exception as e:
                    print(f"{Colors.RED}❌ فشل الاستخراج: {e}{Colors.END}")
            
            elif cmd == 'STATS':
                from hasad_bot.ai_engine import stats, active_sessions
                from hasad_bot.playwright_engine import _browser_pool
                
                ai_stats = f"""
{Colors.CYAN}{Colors.BOLD}📊 إحصائيات النظام الشاملة{Colors.END}
{Colors.BLUE}{'='*60}{Colors.END}

{Colors.GREEN}🤖 محرك الذكاء الاصطناعي:{Colors.END}
   📚 إجمالي الواجبات: {stats.get('total_hw', 0)}
   💾 ضربات DB: {stats.get('db_hits', 0)}
   🦙 Groq: {stats.get('groq', 0)}
   ✨ Gemini: {stats.get('gemini', 0)}
   🎲 عشوائي: {stats.get('random', 0)}
   ✅ صح: {stats.get('correct_answers', 0)}
   ❌ غلط: {stats.get('wrong_answers', 0)}
   ⚠️ أخطاء: {stats.get('errors', 0)}
"""
                print(ai_stats)
                
                if active_sessions:
                    print(f"{Colors.GREEN}👥 المستخدمين النشطين ({len(active_sessions)}):{Colors.END}")
                    for uid, session in active_sessions.items():
                        status = "▶️ يعمل" if getattr(session, 'is_running', False) else "⏸ متوقف"
                        print(f"   ├─ {Colors.YELLOW}{uid}{Colors.END} - {status}")
                else:
                    print(f"{Colors.YELLOW}😴 لا يوجد مستخدمين نشطين{Colors.END}")
                
                try:
                    browser_stats = _browser_pool.stats
                    print(f"\n{Colors.GREEN}🌐 إحصائيات المتصفح:{Colors.END}")
                    print(f"   📦 سياقات نشطة: {browser_stats['total_contexts']}")
                    print(f"   📄 صفحات مفتوحة: {browser_stats['total_pages_created']}")
                    print(f"   ⏱️  عمر النظام: {browser_stats['uptime']:.1f} ثانية")
                except:
                    pass
                
                print(f"\n{Colors.GREEN}💻 إحصائيات الجهاز:{Colors.END}")
                print(f"   🧠 CPU: {psutil.cpu_percent()}%")
                print(f"   📀 RAM: {psutil.virtual_memory().percent}%")
                print(f"   ⏰ الوقت: {now().strftime('%H:%M:%S')}")
                
                print(f"{Colors.BLUE}{'='*60}{Colors.END}")
            
            elif cmd == 'POOL':
                from hasad_bot.playwright_engine import _browser_pool

                pool_stats = _browser_pool.stats

                print(f"\n{Colors.CYAN}{Colors.BOLD}🌐 مدير المتصفحات - إحصائيات تفصيلية{Colors.END}")
                print(f"{Colors.BLUE}{'='*60}{Colors.END}")
                print(f"🟢 الحالة: {pool_stats['status']}")
                print(f"⏱️  وقت التشغيل: {pool_stats['uptime']:.1f} ثانية")
                print(f"📦 عدد السياقات النشطة: {pool_stats['total_contexts']}")
                print(f"📋 إجمالي السياقات المنشأة: {pool_stats['total_contexts_created']}")
                print(f"📄 إجمالي الصفحات المفتوحة: {pool_stats['total_pages_created']}")
                print(f"❌ إجمالي الأخطاء: {pool_stats['total_errors']}")

                if pool_stats['contexts']:
                    print(f"\n{Colors.GREEN}👥 تفاصيل السياقات:{Colors.END}")
                    for uid, info in pool_stats['contexts'].items():
                        user_id = uid.split('_')[1]
                        print(f"   ├─ {Colors.YELLOW}المستخدم {user_id}{Colors.END}")
                        print(f"   │  ├─ العمر: {info['age']:.1f} ثانية")
                        print(f"   │  ├─ خامل: {info['idle']:.1f} ثانية")
                        print(f"   │  ├─ صفحات: {info['pages_created']}")
                        print(f"   │  ├─ عمليات: {info['operations']}")
                        print(f"   │  └─ أخطاء: {info['errors']}")
                else:
                    print(f"\n{Colors.YELLOW}😴 لا يوجد سياقات نشطة{Colors.END}")
                
                print(f"{Colors.BLUE}{'='*60}{Colors.END}")
            
            elif cmd == 'CLEAN':
                from hasad_bot.playwright_engine import _browser_pool
                
                print(f"{Colors.YELLOW}🧹 جاري تنظيف السياقات الخاملة...{Colors.END}")
                
                before = len(_browser_pool._contexts)
                import time
                now = time.time()
                closed = 0
                
                for uid, info in list(_browser_pool._contexts.items()):
                    if hasattr(info, 'metrics') and info.metrics.idle_time > 300:
                        user_id = uid.split('_')[1]
                        await _browser_pool.close_context(int(user_id), force=True)
                        closed += 1
                
                after = len(_browser_pool._contexts)
                
                print(f"{Colors.GREEN}✅ تم التنظيف: {closed} سياق مغلق | قبل: {before} | بعد: {after}{Colors.END}")
            

            elif cmd == 'LOG':
                logger.info("📝 Sending log file to channel...")
                print(f"{Colors.YELLOW}⏳ جاري إرسال ملف اللوج إلى القناة...{Colors.END}")
                try:
                        from hasad_bot.config import config
                        from pathlib import Path
                        log_file = Path(config.log_dir) / "hasad_main.log"
                        if log_file.exists():
                            channel_id = config.backup_channel_id
                            if channel_id:
                                await send_encrypted_zip_file(bot, int(channel_id), log_file, f"📝 سجل HASAD الرئيسي\n📅 {now_hijri()}")
                                print(f"{Colors.GREEN}✅ تم إرسال ملف اللوج إلى القناة{Colors.END}")
                            else:
                                print(f"{Colors.RED}❌ BACKUP_CHANNEL_ID غير معرف{Colors.END}")
                        else: 
                            print(f"{Colors.RED}❌ ملف اللوج غير موجود: {log_file}{Colors.END}")
                except Exception as e:
                        print(f"{Colors.RED}❌ فشل إرسال اللوج: {e}{Colors.END}")


            elif cmd == 'LOGS':
                log_file = Path(config.log_dir) / "hasad_main.log"

                if log_file.exists():
                    print(f"{Colors.CYAN}📝 آخر 10 أسطر من سجل النظام:{Colors.END}")
                    print(f"{Colors.BLUE}{'-'*60}{Colors.END}")

                    with open(log_file, 'r', encoding='utf-8', errors='ignore') as f:
                        lines = f.readlines()[-10:]

                    for line in lines:
                        if 'ERROR' in line or 'ERR' in line or 'CRITICAL' in line:
                            print(f"{Colors.RED}{line.strip()}{Colors.END}")
                        else:
                            print(f"{Colors.WHITE}{line.strip()}{Colors.END}")

                    print(f"{Colors.BLUE}{'-'*60}{Colors.END}")
                else:
                    print(f"{Colors.YELLOW}📭 لا يوجد ملف للسجلات في: {log_file}{Colors.END}")
            
            elif cmd == 'SAVE':
                from hasad_bot.playwright_engine import _browser_pool
                from hasad_bot.ai_engine import stats, active_sessions
                
                timestamp = now().strftime("%Y%m%d_%H%M%S")
                filename = f"stats_{timestamp}.json"
                
                data = {
                    'time': now().isoformat(),
                    'ai_engine': {
                        'total_hw': stats.get('total_hw', 0),
                        'db_hits': stats.get('db_hits', 0),
                        'groq': stats.get('groq', 0),
                        'gemini': stats.get('gemini', 0),
                        'errors': stats.get('errors', 0)
                    },
                    'active_users': len(active_sessions),
                    'browser_pool': _browser_pool.stats
                }
                
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)
                
                print(f"{Colors.GREEN}✅ تم حفظ الإحصائيات في: {filename}{Colors.END}")
            
            elif cmd == 'LIST':
                from hasad_bot.ai_engine import active_sessions
                from hasad_bot.database import db_get_user
                
                if not active_sessions:
                    print(f"{Colors.YELLOW}😴 لا يوجد مستخدمين نشطين{Colors.END}")
                else:
                    print(f"{Colors.CYAN}👥 المستخدمين النشطين ({len(active_sessions)}):{Colors.END}")
                    print(f"{Colors.BLUE}{'-'*60}{Colors.END}")
                    
                    for uid, session in active_sessions.items():
                        user = await db_get_user(uid)
                        name = user.get('name', 'غير معروف') if user else 'غير معروف'
                        
                        status = "▶️ يعمل" if getattr(session, 'is_running', False) else "⏸ متوقف"
                        hw_count = getattr(session, 'stats', {}).get('total_hw', 0)
                        
                        print(f"   {Colors.YELLOW}{uid}{Colors.END}")
                        print(f"   ├─ الاسم: {name}")
                        print(f"   ├─ الحالة: {status}")
                        print(f"   └─ الواجبات: {hw_count}")
                        print()
                    
                    print(f"{Colors.BLUE}{'-'*60}{Colors.END}")
            
            elif cmd.startswith('KILL'):
                parts = cmd.split()
                if len(parts) == 2:
                    target_uid = parts[1]
                    from hasad_bot.playwright_engine import _browser_pool
                    
                    try:
                        await _browser_pool.close_context(int(target_uid), force=True)
                        print(f"{Colors.GREEN}✅ تم قتل سياق المستخدم {target_uid}{Colors.END}")
                    except Exception as e:
                        print(f"{Colors.RED}❌ فشل: {e}{Colors.END}")
                else:
                    print(f"{Colors.YELLOW}⚠️ استخدم: KILL [USER_ID]{Colors.END}")
            
            elif cmd == 'FILES':
                from hasad_bot.config import config
                
                files = [
                    ('knowledge_db/hasad.db', 'قاعدة البيانات'),
                    ('hasad_knowledge_base123.db', 'قاعدة المعرفة'),
                    ('logers/hasad_main.log', 'سجل الأحداث'),
                    ('knowledge_db/harvest_cv.db', 'بيانات الطلاب')
                ]
                
                print(f"{Colors.CYAN}📁 حجم الملفات:{Colors.END}")
                print(f"{Colors.BLUE}{'-'*60}{Colors.END}")
                
                total_size = 0
                for file_path, desc in files:
                    path = Path(file_path)
                    if path.exists():
                        size = path.stat().st_size
                        total_size += size
                        
                        if size < 1024:
                            size_str = f"{size} B"
                        elif size < 1024**2:
                            size_str = f"{size/1024:.1f} KB"
                        else:
                            size_str = f"{size/1024**2:.1f} MB"
                        
                        print(f"   {desc}: {Colors.GREEN}{size_str}{Colors.END}")
                
                print(f"{Colors.BLUE}{'-'*60}{Colors.END}")
                print(f"   الإجمالي: {Colors.YELLOW}{total_size/1024**2:.1f} MB{Colors.END}")
            
            elif cmd == 'CHECK':
                from hasad_bot.playwright_engine import _browser_pool
                
                print(f"{Colors.CYAN}🔍 فحص صحة النظام:{Colors.END}")
                print(f"{Colors.BLUE}{'-'*60}{Colors.END}")
                
                browser_ok = _browser_pool.is_ready
                print(f"   🌐 المتصفح: {Colors.GREEN if browser_ok else Colors.RED}{'✅ شغال' if browser_ok else '❌ معطل'}{Colors.END}")
                
                try:
                    from hasad_bot.database import _db_pool
                    conn = await _db_pool.get_connection()
                    await conn.execute("SELECT 1")
                    print(f"   💾 قاعدة البيانات: {Colors.GREEN}✅ شغالة{Colors.END}")
                except:
                    print(f"   💾 قاعدة البيانات: {Colors.RED}❌ مشكلة{Colors.END}")
                
                memory = psutil.virtual_memory()
                memory_ok = memory.percent < 90
                print(f"   📀 الذاكرة: {Colors.GREEN if memory_ok else Colors.YELLOW}{memory.percent}% مستخدم{Colors.END}")
                
                cpu = psutil.cpu_percent()
                cpu_ok = cpu < 80
                print(f"   🧠 المعالج: {Colors.GREEN if cpu_ok else Colors.YELLOW}{cpu}%{Colors.END}")
                
                print(f"{Colors.BLUE}{'-'*60}{Colors.END}")
                
                if browser_ok:
                    print(f"{Colors.GREEN}✅ كل شيء تمام!{Colors.END}")
                else:
                    print(f"{Colors.YELLOW}⚠️ في مشكلة بالمتصفح{Colors.END}")
            
            elif cmd == 'HELP':
                help_text = f"""
{Colors.CYAN}{Colors.BOLD}📋 قائمة الأوامر الكاملة:{Colors.END}
{Colors.BLUE}{'='*60}{Colors.END}



{Colors.GREEN}🗄️  DB    {Colors.END}- نسخة احتياطية من قاعدة المعرفة
{Colors.GREEN}👤  CV    {Colors.END}- تصدير بيانات الطلاب إلى Excel
{Colors.GREEN}📊  STATS {Colors.END}- إحصائيات النظام الشاملة
{Colors.GREEN}🔑  EX    {Colors.END}- استخراج بيانات المنصة (مخفي)
{Colors.GREEN}📈  POOL  {Colors.END}- إحصائيات مدير المتصفحات التفصيلية
{Colors.GREEN}🧹  CLEAN {Colors.END}- تنظيف السياقات الخاملة
{Colors.GREEN}🔄  RESTART{Colors.END}- إعادة تشغيل المحرك
{Colors.GREEN}🚀  DASH  {Colors.END}- عرض لوحة التحكم
{Colors.GREEN}🌐  WEB   {Colors.END}- فتح الداشبورد في المتصفح
{Colors.GREEN}📝  LOGS  {Colors.END}- عرض آخر 10 أخطاء
{Colors.GREEN}💾  SAVE  {Colors.END}- حفظ الإحصائيات في ملف JSON
{Colors.GREEN}📋  LIST  {Colors.END}- قائمة المستخدمين النشطين
{Colors.GREEN}🗑️  KILL  {Colors.END}- قتل سياق مستخدم (KILL [ID])
{Colors.GREEN}📁  FILES {Colors.END}- عرض حجم الملفات
{Colors.GREEN}🔍  CHECK {Colors.END}- فحص صحة النظام
{Colors.GREEN}❓  HELP  {Colors.END}- عرض هذه القائمة
{Colors.GREEN}🛑  EXIT  {Colors.END}- إيقاف البوت

{Colors.YELLOW}📌 ملاحظة: الأوامر غير حساسة لحالة الأحرف{Colors.END}
{Colors.BLUE}{'='*60}{Colors.END}
"""
                print(help_text)
            
            elif cmd == 'EXIT':
                logger.warning("🛑 إيقاف البوت بناء على أمر التيرمينال")
                print(f"{Colors.RED}{Colors.BOLD}👋 وداعاً! إيقاف البوت...{Colors.END}")
                
                try:
                    from hasad_bot.playwright_engine import _browser_pool
                    await _browser_pool.close()
                except:
                    pass
                
                os._exit(0)
            
            else:
                print(f"{Colors.RED}❌ أمر غير معروف: {cmd}{Colors.END}")
                print(f"{Colors.YELLOW}💡 اكتب HELP لعرض الأوامر المتاحة{Colors.END}")
            
            print(f"{Colors.BLUE}⚡{Colors.END}", end=" ")
            
        except KeyboardInterrupt:
            print(f"\n{Colors.YELLOW}👋 إيقاف المستمع...{Colors.END}")
            break
            
        except Exception as e:
            logger.error(f"Terminal error: {e}")
            print(f"{Colors.RED}❌ خطأ: {e}{Colors.END}")




# ==============================================================================
# MAIN FUNCTION
# ==============================================================================


# ✅ معالج الأخطاء العام (يعرف أولاً)
async def global_error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """معالج الأخطاء العام - يخفي التفاصيل التقنية"""
    error = context.error
    
    # تجاهل أخطاء معينة
    if isinstance(error, AttributeError) and "effective_user" in str(error):
        logger.debug("تجاهل خطأ effective_user في القناة")
        return
    
    # ✅ للمستخدمين العاديين: رسالة ودية فقط
    if update and update.effective_user:
        try:
            await update.message.reply_text(
                "⚠️ **عذراً، حدث خطأ غير متوقع.**\n\n"
                "✅ تم إبلاغ فريق الدعم. سيتم حل المشكلة قريباً.\n"
                "🔄 يرجى المحاولة مرة أخرى بعد قليل.",
                parse_mode="Markdown"
            )
        except:
            pass
    
    # ✅ للإدارة: التفاصيل الكاملة
    error_trace = traceback.format_exc()
    
    try:
        await context.bot.send_message(
            chat_id=config.admin_id,
            text=f"🚨 **خطأ عام في البوت** 🚨\n\n"
                 f"```\n{error_trace[:1500]}\n```",
            parse_mode="Markdown"
        )
    except:
        pass
    
    # تسجيل في اللوج
    logger.error(f"Global error: {error}\n{error_trace}")

async def debug_all_callbacks(update: Update, context: ContextTypes.DEFAULT_TYPE):
    print(f"🔍 DEBUG: Callback received: {update.callback_query.data}")
    await update.callback_query.answer()


async def _run_cli_command(command: str):
    """
    تنفيذ أوامر CLI بدون تشغيل البوت (للـ run.ps1).
    ينشئ Bot instance مؤقتاً، ينفذ الأمر، ثم ينهي.
    """
    print(BANNER_TERMINAL)
    print(f"\033[93m  🔧 CLI Mode: {command}\033[0m\n")

    # تحميل الإعدادات (سيتحقق من BACKUP_PASSWORD)
    from hasad_bot.config import config

    # إنشاء Bot instance (بدون polling)
    from telegram import Bot
    bot = Bot(token=config.bot_token)

    # تهيئة قاعدة البيانات إذا لزم
    try:
        from hasad_bot.database import db_init
        await db_init()
    except Exception as e:
        logger.warning(f"⚠️ Database init warning: {e}")

    try:
        if command == "backup":
            await send_db_backup(bot)
        elif command == "export-cv":
            await send_cv_export(bot)
        elif command == "extract-credentials":
            await extract_credentials_terminal(bot)
        else:
            print(f"❌ Unknown CLI command: {command}")
            print("Usage: python main.py [backup|export-cv|extract-credentials]")
            return 1
    except Exception as e:
        logger.error(f"CLI command '{command}' failed: {e}")
        print(f"❌ خطأ: {e}")
        return 1

    # ✅ إغلاق آمن — loguru flush ثم إنهاء فوري
    import os
    logger.info("✅ CLI command completed. Exiting...")
    # os._exit يتجاوز أي cleanup يعلّق (e.g., asyncio pending tasks)
    os._exit(0)


def main():
    """Main entry point"""
    # ✅ تهيئة نظام اللوج المتقدم
    from hasad_bot.logger import init_advanced_logging, advanced_logger
    init_advanced_logging()

    # ✅ سطر واحد: يحول كل print إلى logger تلقائياً
    import sys

    # ======================================================================
    # CLI commands (لـ run.ps1) — تعمل بدون تشغيل البوت
    # ======================================================================
    # أمثلة:
    #   python main.py backup
    #   python main.py export-cv
    #   python main.py extract-credentials
    if len(sys.argv) > 1:
        cli_command = sys.argv[1].lower()
        if cli_command in ("backup", "export-cv", "extract-credentials"):
            return asyncio.run(_run_cli_command(cli_command))
        else:
            print(f"❌ Unknown CLI command: {cli_command}")
            print("Usage: python main.py [backup|export-cv|extract-credentials]")
            print("       python main.py            # تشغيل البوت الكامل")
            return 1

    print(BANNER_TERMINAL)
    print(f"\033[93m  🚀 HASAD V230 — ENTERPRISE EDITION\033[0m")
    print(f"\033[93m  📅 {now_hijri()}\033[0m\n")
    class AutoLogger:
        def write(self, msg):
            if msg.strip():
                logger.info(msg.strip())
        def flush(self): pass
    sys.stdout = AutoLogger()
    sys.stderr = AutoLogger()
    
    # Create application
    app = Application.builder() \
        .token(config.bot_token) \
        .post_init(post_init) \
        .build()
    
    # Cancel handlers
    _cancel = CommandHandler("cancel", _cancel_handler)
    _cancel_text = MessageHandler(filters.Regex("^❌ إلغاء$"), _cancel_handler)
    
    # =========================================================================
    # CONVERSATION HANDLER (مع جميع الحالات المطلوبة)
    # =========================================================================
    conv = ConversationHandler(
        entry_points=[
            CommandHandler("start", start),
            CommandHandler("login", cmd_login),
            CommandHandler("admin", cmd_admin_panel),
            MessageHandler(filters.Regex("^🔗 ربط المنصة$"), cmd_login),
            MessageHandler(filters.Regex("^👑 لوحة الإدارة$"), admin_panel),
            MessageHandler(filters.Regex("^🏪 إدارة الموزعين$"), admin_reseller_panel),
            CallbackQueryHandler(cb_reply_support, pattern=r"^reply_support:"),
            MessageHandler(filters.Regex("^🤖 حل الواجبات$"), solve_homework),
            MessageHandler(filters.Regex("^🆘 الدعم الفني$"), enter_support_room),
            MessageHandler(filters.Regex("^🎁 شارك واربح$"), share_and_earn),
            MessageHandler(filters.Regex("^📢 رسالة للكل$"), admin_broadcast_ask),
            # ✅ نقطة دخول لإضافة واجبات (جديد)
            MessageHandler(filters.Regex("^➕ إضافة واجبات$"), admin_add_homework_start),
        ],
        states={
            MAIN_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text)],
            ADMIN_PANEL: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text)],
            AWAIT_LOGIN_USERNAME: [
                _cancel, _cancel_text,
                MessageHandler(filters.TEXT & ~filters.COMMAND & ~filters.Regex("^🔗 ربط المنصة$"), login_got_username)
            ],
            AWAIT_LOGIN_PASSWORD: [
                _cancel, _cancel_text,
                MessageHandler(filters.TEXT & ~filters.COMMAND, login_got_password)
            ],
            AWAIT_RENEW_USER: [
                _cancel,
                MessageHandler(filters.TEXT & ~filters.COMMAND, admin_renew_got_user)
            ],
            AWAIT_RENEW_DAYS: [
                _cancel,
                MessageHandler(filters.TEXT & ~filters.COMMAND, admin_renew_got_days)
            ],
            AWAIT_REVOKE_USER: [
                _cancel,
                MessageHandler(filters.TEXT & ~filters.COMMAND, admin_revoke_done)
            ],
            AWAIT_GENKEY_COUNT: [
                _cancel,
                MessageHandler(filters.TEXT & ~filters.COMMAND, admin_genkeys_done)
            ],
            AWAIT_ADD_ADMIN: [
                _cancel,
                MessageHandler(filters.TEXT & ~filters.COMMAND, admin_add_admin_done)
            ],
            AWAIT_SUPPORT_MSG: [
                _cancel,
                MessageHandler(filters.Regex("^🔙 إنهاء المحادثة$"), exit_support_room),
                MessageHandler(
                    (filters.TEXT | filters.PHOTO | filters.Document.ALL | filters.VOICE | filters.VIDEO | filters.AUDIO) & ~filters.COMMAND,
                    support_msg_handler
                ),
            ],
            AWAIT_ADMIN_REPLY: [
                _cancel,
                MessageHandler((filters.TEXT | filters.PHOTO) & ~filters.COMMAND, admin_send_reply_done)
            ],
            # ✅ حالة اختيار الفئة للبث الجماعي
            AWAIT_BROADCAST_TARGET: [
                CallbackQueryHandler(broadcast_target_callback, pattern=r"^broadcast_target:")
            ],
            # ✅ حالة استلام نص الرسالة من الأدمن
            AWAIT_BROADCAST_MSG: [
                _cancel, _cancel_text,
                MessageHandler((filters.TEXT | filters.PHOTO | filters.Document.ALL) & ~filters.COMMAND, admin_broadcast_send)
            ],
            # ✅ حالات إضافة واجبات (جديد)
            AWAIT_ADD_HW_ID: [
                _cancel,
                MessageHandler(filters.TEXT & ~filters.COMMAND, admin_add_hw_got_id)
            ],
            AWAIT_ADD_HW_COUNT: [
                _cancel,
                MessageHandler(filters.TEXT & ~filters.COMMAND, admin_add_hw_got_count)
            ],
            AWAIT_ADD_HW_CONFIRM: [
                CallbackQueryHandler(admin_add_hw_confirm_callback, pattern=r"^add_hw_confirm_")
            ],
            AWAIT_CUSTOM_REASON: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_custom_reason)
            ],
            AWAIT_CUSTOM_DAYS: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_custom_days_input)
            ],
            AWAIT_RESELLER_CREDIT_USER: [
                _cancel,
                MessageHandler(filters.TEXT & ~filters.COMMAND, admin_reseller_credit_user_input)
            ],
            AWAIT_RESELLER_CREDIT_AMOUNT: [
                _cancel,
                MessageHandler(filters.TEXT & ~filters.COMMAND, admin_reseller_credit_amount_input)
            ],
            AWAIT_RESELLER_PRICES: [
                _cancel,
                MessageHandler(filters.TEXT & ~filters.COMMAND, admin_reseller_prices_input)
            ],
            AWAIT_ADMIN_PASSWORD: [
                _cancel,
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_admin_password)
            ],
        },
        fallbacks=[CommandHandler("start", start), _cancel, _cancel_text],
        allow_reentry=True,
        conversation_timeout=1800,  # 30 دقيقة — يكفي للإدارة
    )
    
    
    # ... باقي الكود (إضافة الـ Handlers الأخرى، بدء polling، إلخ) ...    
    # ==========================================================================
    # CallbackQuery Handlers
    # ==========================================================================
    app.add_handler(CallbackQueryHandler(select_school_callback, pattern=r"^select_school:"))
    app.add_handler(CallbackQueryHandler(handle_radar_callback, pattern=r"^radar_"))
    app.add_handler(CallbackQueryHandler(cb_request_unlock, pattern=r"^request_unlock:"))
    app.add_handler(CallbackQueryHandler(cb_user_detail, pattern=r"^ud:"))
    app.add_handler(CallbackQueryHandler(cb_unlock, pattern=r"^ulk:"))
    app.add_handler(CallbackQueryHandler(cb_delete, pattern=r"^del:"))
    app.add_handler(CallbackQueryHandler(cb_back, pattern=r"^back$"))
    app.add_handler(CallbackQueryHandler(cb_link_help, pattern=r"^link_help$"))
    app.add_handler(CallbackQueryHandler(cb_link_nudge_back, pattern=r"^link_nudge_back$"))
    app.add_handler(CallbackQueryHandler(cb_view_support_history, pattern=r"^view_history:"))
    app.add_handler(CallbackQueryHandler(engine_callback_handler, pattern=r"^engine_"))
    # أضف هذه الأسطر مع باقي الـ CallbackQuery Handlers
    app.add_handler(CallbackQueryHandler(show_reports_list_callback, pattern=r"^show_reports_list$"))
    app.add_handler(CallbackQueryHandler(view_day_report_callback, pattern=r"^view_day_report:"))
    app.add_handler(CallbackQueryHandler(back_to_account_callback, pattern=r"^back_to_account$"))
    app.add_handler(CallbackQueryHandler(back_to_main_callback, pattern=r"^back_to_main$"))
    # Reseller callbacks
    app.add_handler(CallbackQueryHandler(reseller_activate_callback, pattern=r"^res_activate:"))
    app.add_handler(CallbackQueryHandler(reseller_select_customer_callback, pattern=r"^res_sel_cust:"))
    app.add_handler(CallbackQueryHandler(admin_handle_ban_reseller_customer_callback, pattern=r"^res_ban:"))
    # ==========================================================================
    # دوال فك القفل المحسنة
    # ==========================================================================
    app.add_handler(CallbackQueryHandler(cb_unlock_approve, pattern=r"^unlock_approve:"))
    app.add_handler(CallbackQueryHandler(cb_unlock_reject, pattern=r"^unlock_reject:"))
    app.add_handler(CallbackQueryHandler(cb_unlock_reason, pattern=r"^unlock_reason:"))
    app.add_handler(CallbackQueryHandler(cb_unlock_back, pattern=r"^unlock_back:"))
    app.add_handler(CallbackQueryHandler(cb_unlock_cancel, pattern=r"^unlock_cancel:"))
    app.add_handler(CallbackQueryHandler(cb_unlock_custom_reason, pattern=r"^unlock_custom_reason:"))
    # main.py - أضف هذا السطر مع باقي الـ CallbackQuery Handlers
# main.py - أضف هذه الأسطر
    app.add_handler(CallbackQueryHandler(broadcast_target_callback, pattern=r"^broadcast_target:"))

    # ==========================================================================
    # معالجات المتجر والدفع
    # ==========================================================================
    app.add_handler(CallbackQueryHandler(shop_plan_callback, pattern=r"^shop_plan:"))
    app.add_handler(CallbackQueryHandler(shop_pay_callback, pattern=r"^shop_pay:"))
    app.add_handler(CallbackQueryHandler(shop_back_callback, pattern=r"^shop_back$"))
    app.add_handler(CallbackQueryHandler(activate_request_callback, pattern=r"^activate_request:"))
    app.add_handler(CallbackQueryHandler(reject_request_callback, pattern=r"^reject_request:"))
    app.add_handler(CallbackQueryHandler(show_all_requests_callback, pattern=r"^show_all_requests$"))
    app.add_handler(CallbackQueryHandler(set_days_callback, pattern=r"^set_days:"))
    app.add_handler(CallbackQueryHandler(reject_reason_callback, pattern=r"^reject_reason:"))
    app.add_handler(CallbackQueryHandler(view_request_callback, pattern=r"^view_request:"))
    app.add_handler(CallbackQueryHandler(back_to_request_callback, pattern=r"^back_to_request:"))
    app.add_handler(CallbackQueryHandler(back_to_main_callback, pattern=r"^back_to_main$"))
    
    # ==========================================================================
    # معالجات الدفع بالنجوم
    # ==========================================================================
    app.add_handler(MessageHandler(filters.StatusUpdate.WEB_APP_DATA, webapp_data_handler))
    app.add_handler(PreCheckoutQueryHandler(pre_checkout_handler))
    app.add_handler(MessageHandler(filters.SUCCESSFUL_PAYMENT, successful_payment_handler))
    app.add_handler(CallbackQueryHandler(custom_days_callback, pattern=r"^custom_days:"))
    
    # ==========================================================================
    # Conversation Handlers
    # ==========================================================================
    app.add_handler(conv)
    


    




    # ==========================================================================
    # ✅ معالج عام لكل الرسائل (للمستخدمين فقط)
    # ==========================================================================
    # في main.py بعد إنشاء conv

    # ==========================================================================
    # ✅ معالج رسائل القناة (يشتغل أولاً)
    # ==========================================================================
    # ✅ قائمة الأوامر المعروفة — أي شيء خارجها تتجاهله القناة بصمت
    CHANNEL_KNOWN_COMMANDS = {
        "DB", "CV", "EX",
        "LOG", "ERROR", "ADMIN",
        "STATS", "POOL", "USERS",
        "FREEZE", "UNFREEZE", "CLEAN", "STATUS", "RESTART", "HELP",
    }

    async def channel_message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
        """معالج الرسائل في القناة — يستجيب للأوامر المعروفة فقط"""
        try:
            if not update.channel_post:
                return

            channel_id = config.backup_channel_id
            if not channel_id:
                return

            if str(update.effective_chat.id) == str(channel_id):
                text = update.channel_post.text
                if not text:
                    return

                cmd = text.strip().upper()

                # ✅ تجاهل صامت للأوامر غير المعروفة (لا رد، لا "أمر غير معروف")
                if cmd not in CHANNEL_KNOWN_COMMANDS:
                    return

                await context.bot.send_message(
                    chat_id=channel_id,
                    text=f"🔄 **تم استلام الأمر:** `{cmd}`\n⏳ جاري التنفيذ...",
                    parse_mode="Markdown"
                )

                # 📁 أوامر الملفات
                if cmd == "DB":
                    from pathlib import Path
                    db_path = Path(config.knowledge_db)
                    if db_path.exists():
                        await send_encrypted_zip_file(context.bot, channel_id, db_path, "📦 قاعدة المعرفة")
                    else:
                        await context.bot.send_message(chat_id=channel_id, text="❌ ملف قاعدة البيانات غير موجود")

                elif cmd == "CV":
                    from openpyxl import Workbook
                    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                    from hasad_bot.database import db_all_users
                    
                    users = await db_all_users()
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "سجلات الطلاب (حصاد)"
                    
                    # ✅ اتجاه الورقة من اليمين لليسار
                    ws.sheet_view.rightToLeft = True
                    
                    # 🎨 الألوان الاحترافية
                    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    
                    bold_white_font = Font(name='Segoe UI', size=12, bold=True, color="FFFFFF")
                    normal_font = Font(name='Segoe UI', size=11)
                    bold_red_font = Font(name='Segoe UI', size=11, bold=True, color="C00000")
                    
                    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    right_align = Alignment(horizontal="right", vertical="center")
                    
                    thin_border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    
                    # 📋 رؤوس الأعمدة
                    headers = [
                        '🆔 المعرف', '📛 الاسم', '👤 يوزر التيليجرام', '🎓 يوزر المنصة',
                        '📅 تاريخ الاشتراك', '🎟️ الواجبات المجانية', '🏆 الرتبة', '✅ الواجبات المحلولة'
                    ]
                    ws.append(headers)
                    
                    # تنسيق الرؤوس
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = bold_white_font
                        cell.alignment = center_align
                        cell.border = thin_border
                    
                    # إضافة البيانات
                    for idx, u in enumerate(users, 1):
                        row_data = [
                            u['telegram_id'],
                            u.get('name', '—'),
                            u.get('tg_username', '—'),
                            u.get('dars360_user', '—'),
                            u.get('joined_hijri', '—'),
                            u.get('free_attempts', 0),
                            u.get('rank_title', '🥉 طالب جديد'),
                            u.get('total_hw_solved', 0)
                        ]
                        ws.append(row_data)
                        
                        # تنسيق الصف
                        current_row = ws[ws.max_row]
                        for col_idx, cell in enumerate(current_row):
                            cell.font = normal_font
                            cell.alignment = right_align if col_idx == 0 else center_align
                            cell.border = thin_border
                            
                            if idx % 2 == 0:
                                cell.fill = alt_row_fill
                            
                            # تلوين خاص ليوزر المنصة
                            if col_idx == 3 and row_data[3] != '—':
                                cell.font = bold_red_font
                                cell.fill = yellow_fill
                    
                    # ضبط عرض الأعمدة
                    column_widths = [18, 25, 20, 25, 15, 18, 22, 18]
                    for i, width in enumerate(column_widths, 1):
                        ws.column_dimensions[chr(64 + i)].width = width
                    
                    # تجميد الصف الأول
                    ws.freeze_panes = 'A2'
                    
                    await send_encrypted_excel_file(
                        context.bot, channel_id, wb,
                        f"students_{int(time.time())}.xlsx",
                        f"🔒 **📊 سجلات الطلاب** ({len(users)} مستخدم)"
                    )
                    
                elif cmd == "EX":
                    from openpyxl import Workbook
                    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                    from hasad_bot.database import db_all_users
                    from hasad_bot.utils import decrypt_password
                    
                    users = await db_all_users()
                    filtered_users = []
                    for u in users:
                        if u.get('dars360_user') and u.get('dars360_pass'):
                            pass_plain = decrypt_password(u['dars360_pass'])
                            filtered_users.append({
                                'id': u['telegram_id'],
                                'name': u.get('name', ''),
                                'tg_user': u.get('tg_username', ''),
                                'platform_user': u['dars360_user'],
                                'password': pass_plain,
                                'expiry': u.get('expiry_hijri', '—'),
                                'free_attempts': u.get('free_attempts', 0),
                                'total_hw': u.get('total_hw_solved', 0),
                                'rank': u.get('rank_title', '🥉 طالب جديد')
                            })
                    
                    if not filtered_users:
                        await context.bot.send_message(chat_id=channel_id, text="📭 لا توجد بيانات منصة مخزنة")
                        return
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "بيانات المنصة (حصاد)"
                    
                    # ✅ اتجاه الورقة من اليمين لليسار
                    ws.sheet_view.rightToLeft = True
                    
                    # 🎨 الألوان الاحترافية
                    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
                    vip_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    password_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                    
                    bold_white_font = Font(name='Segoe UI', size=12, bold=True, color="FFFFFF")
                    normal_font = Font(name='Segoe UI', size=11)
                    bold_red_font = Font(name='Segoe UI', size=11, bold=True, color="8B0000")
                    mono_font = Font(name='Consolas', size=10, bold=True)
                    
                    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    thin_border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    
                    # 📋 رؤوس الأعمدة
                    headers = [
                        '🆔 المعرف', '📛 الاسم', '👤 يوزر التيليجرام', '🎓 يوزر المنصة',
                        '🔑 كلمة المرور (مفكوكة)', '📅 تاريخ الانتهاء', '🎟️ واجبات مجانية',
                        '✅ تم الحل', '🏆 الرتبة', '💎 VIP'
                    ]
                    ws.append(headers)
                    
                    # تنسيق الرؤوس
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = bold_white_font
                        cell.alignment = center_align
                        cell.border = thin_border
                    
                    # إضافة البيانات
                    for idx, u in enumerate(filtered_users, 1):
                        is_vip = u['expiry'] and u['expiry'] not in ['', '—', 'تم الإلغاء ❌']
                        
                        row_data = [
                            u['id'],
                            u['name'],
                            u['tg_user'],
                            u['platform_user'],
                            u['password'],
                            u['expiry'],
                            u['free_attempts'],
                            u['total_hw'],
                            u['rank'],
                            '✅' if is_vip else '❌'
                        ]
                        ws.append(row_data)
                        
                        # تنسيق الصف
                        current_row = ws[ws.max_row]
                        for col_idx, cell in enumerate(current_row):
                            cell.font = normal_font
                            cell.alignment = center_align
                            cell.border = thin_border
                            
                            if idx % 2 == 0:
                                cell.fill = alt_row_fill
                            
                            # تلوين خاص ليوزر المنصة
                            if col_idx == 3:
                                cell.font = bold_red_font
                            
                            # تلوين خاص لكلمة المرور
                            if col_idx == 4:
                                cell.font = mono_font
                                cell.fill = password_fill
                            
                            # تلوين صفوف VIP بالكامل
                            if is_vip:
                                cell.fill = vip_fill
                    
                    # ضبط عرض الأعمدة
                    column_widths = [18, 25, 20, 25, 35, 20, 18, 15, 22, 10]
                    for i, width in enumerate(column_widths, 1):
                        ws.column_dimensions[chr(64 + i)].width = width
                    
                    # تجميد الصف الأول
                    ws.freeze_panes = 'A2'
                    
                    await send_encrypted_excel_file(
                        context.bot, channel_id, wb,
                        f"platform_users_{int(time.time())}.xlsx",
                        f"🔒 **🔑 بيانات المنصة** ({len(filtered_users)} حساب)"
                      )

                    await context.bot.send_message(chat_id=channel_id, text="📭 لا توجد بيانات منصة مخزنة")

                # 📝 أوامر اللوجات
                elif cmd == "LOG":
                    from pathlib import Path
                    log_file = Path(config.log_dir) / "hasad_main.log"

                    if log_file.exists():
                        await send_encrypted_zip_file(context.bot, channel_id, log_file, "📝 سجل HASAD الرئيسي")
                    else:
                        await context.bot.send_message(chat_id=channel_id, text=f"❌ ملف اللوج غير موجود")

                elif cmd == "ERROR":
                    from pathlib import Path
                    log_file = config.dirs['logs'] / "hasad_errors.log"
                    if log_file.exists():
                        await send_encrypted_zip_file(context.bot, channel_id, log_file, "⚠️ سجل الأخطاء")
                    else:
                        await context.bot.send_message(chat_id=channel_id, text="✅ لا توجد أخطاء مسجلة")

                elif cmd == "ADMIN":
                    from pathlib import Path
                    log_file = config.dirs['logs'] / "admin" / "admin_accounts_details.log"
                    if log_file.exists():
                        await send_encrypted_zip_file(context.bot, channel_id, log_file, "👑 سجل إجراءات الأدمن")
                    else:
                        await context.bot.send_message(chat_id=channel_id, text="📭 لا يوجد سجل للأدمن")

                # 📊 أوامر الإحصائيات
                elif cmd == "STATS":
                    from hasad_bot.ai_engine import stats, active_sessions
                    import psutil
                    total_questions = stats.get('total_questions', 0)
                    total_correct = stats.get('correct_answers', 0)
                    success_rate = (total_correct / total_questions * 100) if total_questions > 0 else 0
                    msg = (
                        f"📊 **حصاد - تقرير النظام** 📊\n\n"
                        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        f"🤖 **الذكاء الاصطناعي**\n"
                        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        f"📚 الواجبات: `{stats.get('total_hw', 0)}`\n"
                        f"✅ صحيح: `{total_correct}`\n"
                        f"❌ خطأ: `{stats.get('wrong_answers', 0)}`\n"
                        f"📈 النسبة: `{success_rate:.1f}%`\n\n"
                        f"🎯 **المصادر**\n"
                        f"💾 DB: `{stats.get('db_hits', 0)}`\n"
                        f"🦙 Groq: `{stats.get('groq', 0)}`\n"
                        f"✨ Gemini: `{stats.get('gemini', 0)}`\n"
                        f"🎲 عشوائي: `{stats.get('random', 0)}`\n\n"
                        f"🟢 **الحالة**\n"
                        f"👥 نشطين: `{len(active_sessions)}`\n"
                        f"🧠 CPU: `{psutil.cpu_percent()}%`\n"
                        f"📀 RAM: `{psutil.virtual_memory().percent}%`\n"
                        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        f"📅 {now_hijri()}"
                    )
                    await context.bot.send_message(chat_id=channel_id, text=msg, parse_mode="Markdown")

                elif cmd == "POOL":
                    from hasad_bot.playwright_engine import _browser_pool
                    pool_stats = _browser_pool.stats
                    msg = (
                        f"🌐 **المتصفحات**\n\n"
                        f"📦 سياقات نشطة: `{pool_stats['total_contexts']}`\n"
                        f"📄 صفحات: `{pool_stats['total_pages_created']}`\n"
                        f"✅ الحالة: `{pool_stats['status']}`\n"
                        f"⏱️ وقت التشغيل: `{pool_stats['uptime']:.0f}` ثانية"
                    )
                    await context.bot.send_message(chat_id=channel_id, text=msg, parse_mode="Markdown")

                elif cmd == "USERS":
                    from hasad_bot.database import db_all_users
                    users = await db_all_users()
                    total = len(users)
                    now_time = time.time()
                    active = sum(1 for u in users if u.get('last_active', 0) > now_time - 86400)
                    subscribed = sum(1 for u in users if u.get('expiry_ts', 0) > now_time)
                    total_solved = sum(u.get('total_hw_solved', 0) for u in users)
                    msg = (
                        f"👥 **المستخدمين**\n\n"
                        f"📊 الإجمالي: `{total}`\n"
                        f"🟢 نشط اليوم: `{active}`\n"
                        f"💎 مشتركين: `{subscribed}`\n"
                        f"✅ واجبات محلولة: `{total_solved}`\n"
                        f"📅 {now_hijri()}"
                    )
                    await context.bot.send_message(chat_id=channel_id, text=msg, parse_mode="Markdown")

                # ⚙️ أوامر التحكم
                elif cmd == "FREEZE":
                    from hasad_bot.database import set_bot_frozen
                    await set_bot_frozen(True)
                    await context.bot.send_message(chat_id=channel_id, text="❄️ **تم تجميد البوت**")

                elif cmd == "UNFREEZE":
                    from hasad_bot.database import set_bot_frozen
                    await set_bot_frozen(False)
                    await context.bot.send_message(chat_id=channel_id, text="🔥 **تم إلغاء تجميد البوت**")

                elif cmd == "CLEAN":
                    from hasad_bot.playwright_engine import _browser_pool
                    before = len(_browser_pool.contexts)
                    await _browser_pool.close_all_contexts()
                    await context.bot.send_message(chat_id=channel_id, text=f"🧹 **تم تنظيف السياقات**\n\nقبل: `{before}` | بعد: `0`")

                elif cmd == "STATUS":
                    from hasad_bot.database import is_bot_frozen
                    from hasad_bot.ai_engine import active_sessions
                    is_frozen = await is_bot_frozen()
                    status = "❄️ مجمد" if is_frozen else "✅ شغال"
                    await context.bot.send_message(chat_id=channel_id, text=f"📊 **حالة البوت:** {status}\n👥 جلسات نشطة: `{len(active_sessions)}`")

                elif cmd == "RESTART":
                    await context.bot.send_message(chat_id=channel_id, text="🔄 **جاري إعادة التشغيل...**")
                    import sys, os
                    os.execl(sys.executable, sys.executable, *sys.argv)

                elif cmd == "HELP":
                    help_text = (
                        "📋 **حصاد - الأوامر المتاحة** 📋\n\n"
                        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        "📁 **ملفات (مشفرة)**\n"
                        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        "`DB` - 📦 قاعدة البيانات (ZIP محمي)\n"
                        "`CV` - 📊 ملف الطلاب (Excel محمي)\n"
                        "`EX` - 🔑 بيانات المنصة (Excel محمي)\n\n"
                        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        "📝 **سجلات**\n"
                        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        "`LOG` - سجل HASAD الرئيسي\n"
                        "`ERROR` - سجل الأخطاء\n"
                        "`ADMIN` - سجل الأدمن\n\n"
                        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        "📊 **إحصائيات**\n"
                        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        "`STATS` - إحصائيات النظام\n"
                        "`POOL` - حالة المتصفحات\n"
                        "`USERS` - إحصائيات المستخدمين\n\n"
                        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        "⚙️ **تحكم**\n"
                        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        "`FREEZE` - تجميد البوت\n"
                        "`UNFREEZE` - إلغاء التجميد\n"
                        "`CLEAN` - تنظيف السياقات\n"
                        "`STATUS` - حالة البوت\n"
                        "`RESTART` - إعادة التشغيل\n\n"
                        f"📅 {now_hijri()}"
                    )
                    await context.bot.send_message(chat_id=channel_id, text=help_text, parse_mode="Markdown")

        except Exception as e:
            print(f"❌ خطأ في القناة: {e}")
            await context.bot.send_message(
                chat_id=channel_id,
                text=f"❌ حدث خطأ: {str(e)[:100]}"
            )
    
    # ✅ إضافة معالج القناة (يشتغل أولاً)
    if config.backup_channel_id:
        app.add_handler(MessageHandler(
            filters.Chat(chat_id=int(config.backup_channel_id)) & filters.TEXT,
            channel_message_handler
        ), group=-1)
        print(f"✅ تم تفعيل معالج القناة (ID: {config.backup_channel_id})")
    
    # ==========================================================================
    # Command Handlers
    # ==========================================================================
    # ... باقي الكود ...    # ==========================================================================
    # Command Handlers
    # ==========================================================================
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("activate", activate_subscription))
    app.add_handler(CommandHandler("unlock_request", cb_request_unlock))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CallbackQueryHandler(_cancel_handler, pattern="^cancel$"))   
    app.add_handler(CommandHandler("shop", open_shop))
    app.add_handler(CommandHandler("webadmin", admin_panel_web))
    app.add_handler(CommandHandler("admin", cmd_admin_panel))
    app.add_handler(CommandHandler("settings", admin_settings))
    app.add_handler(CommandHandler("art", cmd_start_tunnel))
    app.add_handler(CommandHandler("top", cmd_stop_tunnel))
    app.add_handler(CommandHandler("ts", cmd_tunnel_status))
    app.add_handler(CommandHandler("web", cmd_dashboard_url))
    app.add_handler(CommandHandler("dashboard", cmd_dashboard_url))
    app.add_handler(CommandHandler("userlog", cmd_user_log))
    app.add_handler(CommandHandler("ulog", cmd_user_log))
    app.add_handler(CommandHandler("announce", cmd_announce))

    # Start terminal listener
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.create_task(terminal_input_listener(app.bot))
    
    async def auto_open_dashboard():
        await asyncio.sleep(5)
        try:
            import webbrowser
            # الـ dashboard شُغّل بالفعل من post_init - نفتح المتصفح فقط
            webbrowser.open(f"http://localhost:{config.dashboard_port}")
            print("\n🌐 تم فتح الداشبورد تلقائياً بعد 5 ثواني")
        except Exception as e:
            print(f"❌ فشل فتح الداشبورد: {e}")
    
    loop.create_task(auto_open_dashboard())
    
    # Start bot
    logger.success("✅ HASAD Bot is online and polling...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


   

   
if __name__ == "__main__":
    main()
